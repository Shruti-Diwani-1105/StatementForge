import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelGenerator:
    """Generates a professionally formatted Excel spreadsheet from parsed statements."""

    @classmethod
    def generate_excel(cls, pdf_path, bank_name, account_holder, period, transactions):
        """
        Creates an Excel file in the same folder as the PDF.
        Returns the absolute Excel file path.
        """
        # Determine output xlsx path
        base, _ = os.path.splitext(pdf_path)
        excel_path = base + ".xlsx"

        # Calculate metrics
        num_transactions = len(transactions)
        total_debit = sum(t["debit"] for t in transactions)
        total_credit = sum(t["credit"] for t in transactions)
        
        opening_balance = 0.0
        closing_balance = 0.0
        if num_transactions > 0:
            # Reconstruct opening balance based on first row
            first = transactions[0]
            opening_balance = first["balance"] + first["debit"] - first["credit"]
            closing_balance = transactions[-1]["balance"]

        wb = Workbook()
        
        # ----------------------------------------------------
        # 1. SUMMARY SHEET
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Times New Roman", size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        
        section_font = Font(name="Times New Roman", size=11, bold=True, color="1E3A8A")
        
        bold_lbl_font = Font(name="Times New Roman", size=11, bold=True, color="334155")
        regular_val_font = Font(name="Times New Roman", size=11, color="0F172A")
        
        currency_val_font = Font(name="Times New Roman", size=11, bold=True, color="0F172A")
        
        light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Blue 50
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Header Title block
        ws_sum.merge_cells("A1:C1")
        title_cell = ws_sum["A1"]
        title_cell.value = "StatementForge Extraction Summary"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 40

        # Data rows definition
        summary_rows = [
            ("", ""), # row 2 empty
            ("Account Details", ""), # row 3 header
            ("Bank Name", bank_name),
            ("Account Holder", account_holder),
            ("Statement Period", period),
            ("", ""), # row 7 empty
            ("Financial Details", ""), # row 8 header
            ("Opening Balance", opening_balance),
            ("Closing Balance", closing_balance),
            ("Total Debits (Withdrawals)", total_debit),
            ("Total Credits (Deposits)", total_credit),
            ("Total Transactions", num_transactions),
            ("Extraction Date", datetime.datetime.now().strftime("%Y-%m-%d %I:%M %p")),
        ]

        # Currency formatting rule
        currency_format = '"₹ "#,##0.00'

        for idx, (label, val) in enumerate(summary_rows, start=2):
            ws_sum.row_dimensions[idx].height = 22
            
            if not label:
                continue
                
            if val == "": # Header sections
                ws_sum.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=3)
                cell = ws_sum.cell(row=idx, column=1)
                cell.value = label
                cell.font = section_font
                cell.fill = light_blue_fill
                cell.alignment = Alignment(vertical="center", indent=1)
                for col in range(1, 4):
                    ws_sum.cell(row=idx, column=col).border = thin_border
            else: # Key value rows
                c_lbl = ws_sum.cell(row=idx, column=1)
                c_lbl.value = label
                c_lbl.font = bold_lbl_font
                c_lbl.alignment = Alignment(vertical="center", indent=1)
                c_lbl.border = thin_border
                
                c_val = ws_sum.cell(row=idx, column=2)
                c_val.value = val
                c_val.border = thin_border
                c_val.alignment = Alignment(vertical="center", indent=1)
                
                if isinstance(val, (int, float)):
                    c_val.font = currency_val_font
                    if label != "Total Transactions":
                        c_val.number_format = currency_format
                        c_val.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c_val.font = regular_val_font

        # Adjust columns for Summary
        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 35
        ws_sum.column_dimensions["C"].width = 15

        # ----------------------------------------------------
        # 2. TRANSACTIONS SHEET
        # ----------------------------------------------------
        ws_tx = wb.create_sheet(title="Transactions")
        ws_tx.views.sheetView[0].showGridLines = True
        
        # Headers
        headers = ["Row", "Date", "Narration", "Debit (Withdrawals)", "Credit (Deposits)", "Balance"]
        
        hdr_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        
        ws_tx.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_tx.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        # Alternate Row Highlight fill
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Slate 50
        
        # Write rows
        for row_idx, tx in enumerate(transactions, start=2):
            ws_tx.row_dimensions[row_idx].height = 20
            
            c_row = ws_tx.cell(row=row_idx, column=1, value=row_idx - 1)
            c_row.alignment = Alignment(horizontal="center", vertical="center")
            
            c_date = ws_tx.cell(row=row_idx, column=2, value=tx["date"])
            c_date.alignment = Alignment(horizontal="center", vertical="center")
            
            c_narr = ws_tx.cell(row=row_idx, column=3, value=tx["narration"])
            c_narr.alignment = Alignment(vertical="center")
            
            c_deb = ws_tx.cell(row=row_idx, column=4, value=tx["debit"])
            c_deb.number_format = currency_format
            c_deb.alignment = Alignment(horizontal="right", vertical="center")
            
            c_cred = ws_tx.cell(row=row_idx, column=5, value=tx["credit"])
            c_cred.number_format = currency_format
            c_cred.alignment = Alignment(horizontal="right", vertical="center")
            
            c_bal = ws_tx.cell(row=row_idx, column=6, value=tx["balance"])
            c_bal.number_format = currency_format
            c_bal.alignment = Alignment(horizontal="right", vertical="center")
            
            # Apply fonts, borders, zebra striping
            for col_idx in range(1, 7):
                cell = ws_tx.cell(row=row_idx, column=col_idx)
                cell.font = regular_val_font
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

        # Freeze Header Row
        ws_tx.freeze_panes = "A2"

        # Apply Auto Filter
        ws_tx.auto_filter.ref = f"A1:F{num_transactions + 1}"

        # Adjust Columns automatically
        for col in ws_tx.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Row index 1 header label
            for cell in col:
                # If cell has a currency format, assume length includes format characters
                val_str = str(cell.value or "")
                if cell.number_format == currency_format and isinstance(cell.value, (int, float)):
                    val_str = f"₹ {cell.value:,.2f}"
                max_len = max(max_len, len(val_str))
                
            ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Save workbook
        wb.save(excel_path)
        return excel_path
