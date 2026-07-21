import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class GSTExcelWriter:
    """Generates professional dual-sheet GST Ledger reports matching the StatementForge aesthetic."""

    @classmethod
    def write_gst_excel(cls, pdf_path: str, bank_name: str, account_holder: str, period: str, gst_ledger: list) -> str:
        """
        Creates an Excel file beside the PDF containing the GST Summary and the GST transaction ledger.
        Returns the absolute Excel file path.
        """
        base, _ = os.path.splitext(pdf_path)
        excel_path = f"{base}_GST_Report.xlsx"
        
        counter = 1
        while os.path.exists(excel_path):
            excel_path = f"{base}_GST_Report_({counter}).xlsx"
            counter += 1

        wb = Workbook()
        font_name = "Segoe UI"
        
        # Fonts
        title_font = Font(name=font_name, size=15, bold=True, color="FFFFFF")
        subtitle_font = Font(name=font_name, size=9, italic=True, color="E2E8F0")
        section_hdr_font = Font(name=font_name, size=11, bold=True, color="B45309") # Warm Amber/Orange
        bold_lbl_font = Font(name=font_name, size=10, bold=True, color="334155")
        regular_val_font = Font(name=font_name, size=10, color="0F172A")
        bold_val_font = Font(name=font_name, size=10, bold=True, color="0F172A")
        disclaimer_font = Font(name=font_name, size=8, italic=True, color="64748B")
        
        # Header and Row Fills
        amber_hdr_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid") # Dark Amber/Gold
        light_amber_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid") # Amber 50
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Slate zebra
        warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light yellow/orange warning row
        
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        currency_format = '₹ #,##0.00'
        percent_format = '0.0%'
        
        # Mask account number
        from services.gst_service import GSTService
        masked_acc = GSTService.mask_account_number(gst_ledger[0].get("account_number") if gst_ledger else "")

        # ----------------------------------------------------
        # SHEET 1: GST SUMMARY
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "GST Summary"
        ws_sum.views.sheetView[0].showGridLines = True
        
        # Setup page margins and print layout (A4 Landscape)
        ws_sum.page_setup.orientation = ws_sum.ORIENTATION_LANDSCAPE
        ws_sum.page_setup.paperSize = ws_sum.PAPERSIZE_A4
        ws_sum.page_setup.fitToPage = True
        ws_sum.page_setup.fitToWidth = 1
        ws_sum.page_setup.fitToHeight = 0

        # Title Block
        ws_sum.merge_cells("A1:C1")
        title_cell = ws_sum["A1"]
        title_cell.value = "AI-Generated GST Reconciliation & Analysis Report"
        title_cell.font = title_font
        title_cell.fill = amber_hdr_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 32

        # Subtitle Block (Disclaimer)
        ws_sum.merge_cells("A2:C2")
        sub_cell = ws_sum["A2"]
        sub_cell.value = "Disclaimer: GST values are estimated based on transaction patterns and are not official tax filings."
        sub_cell.font = subtitle_font
        sub_cell.fill = amber_hdr_fill
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[2].height = 20

        # Build General details and formulas
        summary_rows = [
            ("", ""),
            ("General Details", ""),
            ("Bank Name", bank_name),
            ("Account Holder", account_holder),
            ("Account Number", masked_acc),
            ("Statement Period", period),
            ("Report Export Date", datetime.datetime.now().strftime("%Y-%m-%d %I:%M %p")),
            ("", ""),
            ("GST Reconciliation Summary", ""),
            ("Total GST Paid (ITC claimable)", "=SUMIF('GST Transactions'!M:M, \"Yes\", 'GST Transactions'!L:L)"),
            ("Total GST Collected (Output tax)", "=SUMIF('GST Transactions'!C:C, \"*Credit*\", 'GST Transactions'!L:L)"),
            ("Net GST Payable/Refundable", "=B14-B13"), # Collected - Paid
            ("Total GST-Linked Transactions", "=COUNTA('GST Transactions'!A:A)-1"),
        ]

        for idx, (label, val) in enumerate(summary_rows, start=3):
            ws_sum.row_dimensions[idx].height = 22
            if not label:
                continue
            
            if val == "":  # Header Sections
                ws_sum.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=3)
                cell = ws_sum.cell(row=idx, column=1, value=label)
                cell.font = section_hdr_font
                cell.fill = light_amber_fill
                for col in range(1, 4):
                    ws_sum.cell(row=idx, column=col).border = thin_border
            else:  # Data Rows
                c_lbl = ws_sum.cell(row=idx, column=1, value=label)
                c_lbl.font = bold_lbl_font
                c_lbl.alignment = Alignment(vertical="center", indent=1)
                c_lbl.border = thin_border
                
                c_val = ws_sum.cell(row=idx, column=2, value=val)
                c_val.border = thin_border
                c_val.alignment = Alignment(vertical="center", indent=1)
                
                if str(val).startswith("="):
                    c_val.font = bold_val_font
                    if label != "Total GST-Linked Transactions":
                        c_val.number_format = currency_format
                        c_val.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    c_val.font = regular_val_font

        # Footnote Disclaimer
        disclaimer_row_idx = len(summary_rows) + 5
        ws_sum.merge_cells(start_row=disclaimer_row_idx, start_column=1, end_row=disclaimer_row_idx+2, end_column=3)
        disc_cell = ws_sum.cell(row=disclaimer_row_idx, column=1)
        disc_cell.value = "Disclaimer Footer: This report has been generated automatically by StatementForge using AI-assisted transaction analysis. GST values are estimated based on available bank statement data and transaction patterns. This report should not be used as an official GST filing document. Users should verify all GST values against tax invoices before statutory filing."
        disc_cell.font = disclaimer_font
        disc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws_sum.column_dimensions["A"].width = 32
        ws_sum.column_dimensions["B"].width = 38
        ws_sum.column_dimensions["C"].width = 15

        # ----------------------------------------------------
        # SHEET 2: GST TRANSACTIONS LEDGER
        # ----------------------------------------------------
        ws_led = wb.create_sheet(title="GST Transactions")
        ws_led.views.sheetView[0].showGridLines = True
        
        # Setup page margins and print layout (A4 Landscape)
        ws_led.page_setup.orientation = ws_led.ORIENTATION_LANDSCAPE
        ws_led.page_setup.paperSize = ws_led.PAPERSIZE_A4
        ws_led.page_setup.fitToPage = True
        ws_led.page_setup.fitToWidth = 1
        ws_led.page_setup.fitToHeight = 0
        
        headers = [
            "Date", "Narration", "Category", "Vendor Name", "Vendor GSTIN", "Total Amount", 
            "Taxable Value", "GST Rate", "CGST", "SGST", "IGST", 
            "Total GST", "ITC Eligible", "GSTR-2B Status", "AI Confidence", "Status"
        ]
        
        # Write headers
        ws_led.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_led.cell(row=1, column=col_idx, value=header)
            cell.font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
            cell.fill = amber_hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Write data rows
        for r_idx, tx in enumerate(gst_ledger, start=2):
            ws_led.row_dimensions[r_idx].height = 22
            
            # Format confidence as fraction for percentage formatting
            conf_val = tx.get("confidence", 100.0)
            if isinstance(conf_val, (int, float)):
                conf_val = conf_val / 100.0
                
            row_data = [
                tx.get("date", ""),
                tx.get("narration", ""),
                tx.get("category", ""),
                tx.get("vendor", ""),
                tx.get("gstin", "Unassigned"),
                tx.get("total_amount", 0.0),
                tx.get("base_value", 0.0),
                tx.get("gst_rate", 0.18),
                tx.get("cgst", 0.0),
                tx.get("sgst", 0.0),
                tx.get("igst", 0.0),
                tx.get("total_gst", 0.0),
                tx.get("itc_eligible", "No"),
                tx.get("gstr2b_status", "Not Reconciled"),
                conf_val,
                tx.get("status", "Estimated")
            ]
            
            is_warn = tx.get("confidence", 100.0) < 70 or tx.get("is_duplicate") or tx.get("is_missing_invoice")
            
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_led.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_val_font
                cell.border = thin_border
                
                # Apply row styling (zebra vs warning highlight)
                if is_warn:
                    cell.fill = warning_fill
                elif r_idx % 2 == 0:
                    cell.fill = zebra_fill
                
                # Column alignments and formats
                if c_idx in (6, 7, 9, 10, 11, 12): # Amounts
                    cell.number_format = currency_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx in (8, 15): # Rates / Percentages
                    cell.number_format = percent_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx in (1, 5, 13, 14, 16): # Center details
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")

        # Freeze headers & enable auto-filtering
        ws_led.freeze_panes = "A2"
        ws_led.auto_filter.ref = f"A1:P{len(gst_ledger) + 1}"

        # Adjust columns dynamically
        for col in ws_led.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            c_idx = col[0].column
            for cell in col:
                val_str = str(cell.value or "")
                if cell.value is not None:
                    if c_idx in (6, 7, 9, 10, 11, 12) and isinstance(cell.value, (int, float)):
                        val_str = f"₹ {cell.value:,.2f}"
                    elif c_idx in (8, 15) and isinstance(cell.value, (int, float)):
                        val_str = f"{cell.value * 100:.1f}%"
                max_len = max(max_len, len(val_str))
            ws_led.column_dimensions[col_letter].width = max(max_len + 3, 10)

        wb.save(excel_path)
        return excel_path
