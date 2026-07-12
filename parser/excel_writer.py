import os
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from parser.utils import ParserUtils

class ExcelWriter:
    """Generates professional dual-sheet bank statement Excel spreadsheets."""

    @classmethod
    def write_excel(cls, pdf_path: str, bank_name: str, account_holder: str, period: str, transactions: list) -> str:
        """
        Builds the Excel sheet beside the PDF, handles collisions, styles elements, and returns output path.
        """
        base, _ = os.path.splitext(pdf_path)
        excel_path = base + ".xlsx"
        counter = 1
        while os.path.exists(excel_path):
            excel_path = f"{base}_({counter}).xlsx"
            counter += 1

        wb = Workbook()

        font_name = "Segoe UI"
        title_font = Font(name=font_name, size=14, bold=True, color="FFFFFF")
        section_hdr_font = Font(name=font_name, size=11, bold=True, color="1E3A8A")
        bold_label_font = Font(name=font_name, size=11, bold=True, color="334155")
        regular_val_font = Font(name=font_name, size=11, color="0F172A")
        bold_val_font = Font(name=font_name, size=11, bold=True, color="0F172A")

        blue_hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        light_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        currency_format = '₹ #,##0.00'

        # ----------------------------------------------------
        # SHEET 1: TRANSACTIONS
        # ----------------------------------------------------
        ws_tx = wb.active
        ws_tx.title = "Transactions"
        ws_tx.views.sheetView[0].showGridLines = True

        all_keys = set()
        for tx in transactions:
            all_keys.update(tx.keys())

        preferred_order = [
            ("date", "Date"),
            ("value_date", "Value Date"),
            ("value date", "Value Date"),
            ("narration", "Transaction Description / Narration"),
            ("description", "Transaction Description / Narration"),
            ("ref_no", "Cheque Number / Reference Number"),
            ("ref no", "Cheque Number / Reference Number"),
            ("debit", "Debit"),
            ("credit", "Credit"),
            ("balance", "Balance"),
            ("transaction_type", "Transaction Type"),
            ("type", "Transaction Type")
        ]

        headers = []
        key_order = []
        seen_keys = set()

        for key_pat, header_name in preferred_order:
            for actual_key in all_keys:
                if actual_key.lower() == key_pat and actual_key not in seen_keys:
                    headers.append(header_name)
                    key_order.append(actual_key)
                    seen_keys.add(actual_key)
                    break

        for actual_key in all_keys:
            if actual_key not in seen_keys:
                header_name = actual_key.replace("_", " ").title()
                headers.append(header_name)
                key_order.append(actual_key)
                seen_keys.add(actual_key)

        # Write header
        ws_tx.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_tx.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
            cell.fill = blue_hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Write data rows - strictly ensuring every transaction is output
        for row_idx, tx in enumerate(transactions, start=2):
            ws_tx.row_dimensions[row_idx].height = 20
            for col_idx, key in enumerate(key_order, start=1):
                cell = ws_tx.cell(row=row_idx, column=col_idx)
                val = tx.get(key)
                header_lower = headers[col_idx - 1].lower()

                is_currency = any(term in header_lower for term in ["debit", "credit", "balance", "amount"])
                
                if is_currency:
                    parsed_val = ParserUtils.parse_numeric(val)
                    if parsed_val is not None and isinstance(parsed_val, (int, float)):
                        cell.value = parsed_val
                        cell.number_format = currency_format
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.value = None
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "date" in header_lower:
                    parsed_date = ParserUtils.parse_date(val)
                    cell.value = parsed_date
                    if isinstance(parsed_date, datetime.date):
                        cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    is_text_field = any(term in header_lower for term in ["narration", "description", "particulars", "remarks", "reference", "cheque", "ref no", "type"])
                    if is_text_field:
                        cell.value = str(val) if val is not None else ""
                        cell.number_format = '@'
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        parsed_val = ParserUtils.parse_numeric(val)
                        cell.value = parsed_val
                        if isinstance(parsed_val, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                cell.font = regular_val_font
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

        # Data integrity check: Verify written row count matches input transaction count
        written_rows = ws_tx.max_row - 1 if transactions else 0
        if len(transactions) != written_rows:
            raise ValueError(
                f"Data Integrity Error: Extracted transaction count ({len(transactions)}) "
                f"does not match written Excel row count ({written_rows}). Export aborted."
            )

        ws_tx.freeze_panes = "A2"
        if transactions:
            last_col_letter = get_column_letter(len(headers))
            ws_tx.auto_filter.ref = f"A1:{last_col_letter}{len(transactions) + 1}"


        for col in ws_tx.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    if isinstance(val, datetime.date):
                        val_str = val.strftime('%Y-%m-%d')
                    elif isinstance(val, (int, float)) and any(term in ws_tx.cell(row=1, column=cell.column).value.lower() for term in ["debit", "credit", "balance", "amount"]):
                        val_str = f"₹ {val:,.2f}"
                    else:
                        val_str = str(val)
                    max_len = max(max_len, len(val_str))
            ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # ----------------------------------------------------
        # SHEET 2: SUMMARY
        # ----------------------------------------------------
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.merge_cells("A1:C1")
        title_cell = ws_sum["A1"]
        title_cell.value = "StatementForge Summary Report"
        title_cell.font = title_font
        title_cell.fill = blue_hdr_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 40

        opening_balance = 0.0
        closing_balance = 0.0
        total_debit = 0.0
        total_credit = 0.0

        for tx in transactions:
            deb_val = ParserUtils.parse_numeric(tx.get("debit"))
            if isinstance(deb_val, (int, float)):
                total_debit += deb_val
            cred_val = ParserUtils.parse_numeric(tx.get("credit"))
            if isinstance(cred_val, (int, float)):
                total_credit += cred_val

        if transactions:
            first_tx = transactions[0]
            last_tx = transactions[-1]
            first_bal = ParserUtils.parse_numeric(first_tx.get("balance"))
            last_bal = ParserUtils.parse_numeric(last_tx.get("balance"))
            first_deb = ParserUtils.parse_numeric(first_tx.get("debit")) or 0.0
            first_cred = ParserUtils.parse_numeric(first_tx.get("credit")) or 0.0

            if first_bal is not None:
                opening_balance = first_bal + first_deb - first_cred
            if last_bal is not None:
                closing_balance = last_bal

        summary_rows = [
            ("", ""),
            ("Account Details", ""),
            ("Bank Name", bank_name or "Unknown Bank"),
            ("Account Holder", account_holder or "Unknown"),
            ("Statement Period", period or "Unknown Period"),
            ("", ""),
            ("Financial Details", ""),
            ("Opening Balance", opening_balance),
            ("Closing Balance", closing_balance),
            ("Total Debit", total_debit),
            ("Total Credit", total_credit),
            ("Transaction Count", len(transactions)),
            ("", ""),
            ("Metadata", ""),
            ("Generated Date", datetime.datetime.now().strftime("%Y-%m-%d %I:%M %p")),
            ("Generated By", "StatementForge")
        ]

        for idx, (label, val) in enumerate(summary_rows, start=2):
            ws_sum.row_dimensions[idx].height = 22
            if not label:
                continue

            if val == "":
                ws_sum.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=3)
                cell = ws_sum.cell(row=idx, column=1)
                cell.value = label
                cell.font = section_hdr_font
                cell.fill = light_blue_fill
                cell.alignment = Alignment(vertical="center", indent=1)
                for col in range(1, 4):
                    ws_sum.cell(row=idx, column=col).border = thin_border
            else:
                c_lbl = ws_sum.cell(row=idx, column=1)
                c_lbl.value = label
                c_lbl.font = bold_label_font
                c_lbl.alignment = Alignment(vertical="center", indent=1)
                c_lbl.border = thin_border

                c_val = ws_sum.cell(row=idx, column=2)
                c_val.value = val
                c_val.border = thin_border
                c_val.alignment = Alignment(vertical="center", indent=1)

                if isinstance(val, (int, float)):
                    c_val.font = bold_val_font
                    if label != "Transaction Count":
                        c_val.number_format = currency_format
                        c_val.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        c_val.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c_val.font = regular_val_font

        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 38
        ws_sum.column_dimensions["C"].width = 15

        wb.save(excel_path)
        print("Excel Generated Successfully")
        return excel_path
