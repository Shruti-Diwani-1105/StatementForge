import os
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton,
    QFileDialog, QMessageBox, QGraphicsDropShadowEffect, QTableWidget,
    QTableWidgetItem, QComboBox, QHeaderView, QLineEdit, QSpacerItem, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread
from PyQt6.QtGui import QCursor, QColor

from widgets.custom_button import PrimaryButton, SecondaryButton
from settings.toast import Toast
from services.mongodb_service import MongoDBService
from utils.user_session import UserSession
from services.tally_service import TallyService


class DBQueryWorker(QThread):
    result_ready = pyqtSignal(object)

    def __init__(self, query_fn, parent=None):
        super().__init__(parent)
        self.query_fn = query_fn

    def run(self):
        try:
            res = self.query_fn()
            self.result_ready.emit(res)
        except Exception as e:
            print(f"TallyExport DBQueryWorker error: {e}")
            self.result_ready.emit(None)


class TallyExportWidget(QWidget):
    """
    Renders an interactive Tally Integration screen to map and sync/export statement transactions
    to Tally ERP 9 and Tally Prime (XML, Excel, and Direct HTTP sync).
    """
    closed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_theme = "light"
        
        # State Data
        self.transactions = []
        self.active_excel_path = None
        self.bank_name = "Bank Account"
        self._is_updating_table = False
        
        self.init_ui()
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(500, self.load_statements_dropdown)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(24, 24, 24, 24)
        main_layout.setSpacing(16)

        # ==========================================
        # TOP BAR: TITLE & DESCRIPTION
        # ==========================================
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)
        self.title_lbl = QLabel("Tally Accounting Integration Hub")
        self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #0F172A;")
        self.subtitle_lbl = QLabel("Assign ledger codes, review voucher classifications, and direct sync or export voucher files.")
        self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #64748B;")
        
        text_layout.addWidget(self.title_lbl)
        text_layout.addWidget(self.subtitle_lbl)
        header_layout.addLayout(text_layout)
        header_layout.addStretch()
        
        # Close / Return button
        self.close_btn = SecondaryButton("Back to Dashboard")
        self.close_btn.setFixedWidth(150)
        self.close_btn.clicked.connect(self.close_screen)
        header_layout.addWidget(self.close_btn)
        
        main_layout.addWidget(header_widget)

        # ==========================================
        # CONTROL BAR: LEDGERS, FILTERS & EXPORTS
        # ==========================================
        control_card = QFrame()
        control_card.setObjectName("ControlCard")
        control_card.setStyleSheet("""
            QFrame#ControlCard {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 15))
        shadow.setOffset(0, 2)
        control_card.setGraphicsEffect(shadow)

        control_layout = QHBoxLayout(control_card)
        control_layout.setContentsMargins(16, 16, 16, 16)
        control_layout.setSpacing(16)

        # 1. Statement Selector
        sel_layout = QVBoxLayout()
        sel_layout.setSpacing(4)
        sel_lbl = QLabel("Select Statement:")
        sel_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        self.statement_combo = QComboBox()
        self.statement_combo.setMinimumWidth(220)
        self.statement_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 6px 10px;
                font-weight: 600;
                color: #1E293B;
                background-color: #F8FAFC;
            }
        """)
        self.statement_combo.currentTextChanged.connect(self.on_statement_combo_changed)
        sel_layout.addWidget(sel_lbl)
        sel_layout.addWidget(self.statement_combo)
        control_layout.addLayout(sel_layout)

        # 2. Bank Ledger Input
        bank_ledger_layout = QVBoxLayout()
        bank_ledger_layout.setSpacing(4)
        bank_lbl = QLabel("Tally Bank Ledger:")
        bank_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        self.bank_ledger_input = QLineEdit("Bank Account")
        self.bank_ledger_input.setPlaceholderText("e.g. HDFC Bank Account")
        self.bank_ledger_input.setMinimumWidth(130)
        self.bank_ledger_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 6px 10px;
                color: #1E293B;
                background-color: #FFFFFF;
            }
        """)
        self.bank_ledger_input.textChanged.connect(self.on_global_ledger_changed)
        bank_ledger_layout.addWidget(bank_lbl)
        bank_ledger_layout.addWidget(self.bank_ledger_input)
        control_layout.addLayout(bank_ledger_layout)

        # 3. Default Counter Ledger Input
        counter_layout = QVBoxLayout()
        counter_layout.setSpacing(4)
        counter_lbl = QLabel("Default Counter Ledger:")
        counter_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        self.counter_ledger_input = QLineEdit("Suspense Account")
        self.counter_ledger_input.setPlaceholderText("e.g. Suspense Account")
        self.counter_ledger_input.setMinimumWidth(130)
        self.counter_ledger_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 6px 10px;
                color: #1E293B;
                background-color: #FFFFFF;
            }
        """)
        self.counter_ledger_input.textChanged.connect(self.on_global_ledger_changed)
        counter_layout.addWidget(counter_lbl)
        counter_layout.addWidget(self.counter_ledger_input)
        control_layout.addLayout(counter_layout)

        # 4. Tally Server URL
        server_layout = QVBoxLayout()
        server_layout.setSpacing(4)
        server_lbl = QLabel("Tally Server URL:")
        server_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        self.server_url_input = QLineEdit("http://localhost:9000")
        self.server_url_input.setPlaceholderText("http://localhost:9000")
        self.server_url_input.setMinimumWidth(150)
        self.server_url_input.setStyleSheet("""
            QLineEdit {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 6px 10px;
                color: #1E293B;
                background-color: #FFFFFF;
            }
        """)
        server_layout.addWidget(server_lbl)
        server_layout.addWidget(self.server_url_input)
        control_layout.addLayout(server_layout)

        control_layout.addStretch()

        # 5. Action Sync & Export Buttons
        btn_box = QHBoxLayout()
        btn_box.setSpacing(8)
        
        # Primary Action: Direct Sync to Tally
        self.sync_tally_btn = QPushButton("Sync to Tally")
        self.sync_tally_btn.setToolTip("Directly push and create ledgers & vouchers in your running Tally application.")
        self.sync_tally_btn.setFixedWidth(110)
        self.sync_tally_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.sync_tally_btn.setStyleSheet("""
            QPushButton {
                background-color: #E0F2FE;
                color: #0369A1;
                font-weight: 700;
                font-size: 12px;
                border: 1px solid #BAE6FD;
                border-radius: 6px;
                padding: 8px 12px;
            }
            QPushButton:hover {
                background-color: #BAE6FD;
            }
        """)
        self.sync_tally_btn.clicked.connect(self.sync_direct_to_tally_action)
        btn_box.addWidget(self.sync_tally_btn)

        self.export_xml_btn = PrimaryButton("Export XML")
        self.export_xml_btn.setToolTip("Export XML vouchers file for manual Tally import.")
        self.export_xml_btn.setFixedWidth(100)
        self.export_xml_btn.clicked.connect(self.export_tally_xml_action)
        btn_box.addWidget(self.export_xml_btn)

        self.export_excel_btn = QPushButton("Export Excel")
        self.export_excel_btn.setToolTip("Export structured Excel sheet for Tally Prime import mapping.")
        self.export_excel_btn.setFixedWidth(100)
        self.export_excel_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #F0FDF4;
                color: #16A34A;
                font-weight: 600;
                font-size: 12px;
                border: 1px solid #BBF7D0;
                border-radius: 6px;
                padding: 8px 12px;
            }
            QPushButton:hover {
                background-color: #DCFCE7;
            }
        """)
        self.export_excel_btn.clicked.connect(self.export_tally_excel_action)
        btn_box.addWidget(self.export_excel_btn)

        control_layout.addLayout(btn_box)
        main_layout.addWidget(control_card)

        # ==========================================
        # TABLE PREVIEW CONTAINER
        # ==========================================
        grid_container = QFrame()
        grid_container.setObjectName("GridContainer")
        grid_container.setStyleSheet("""
            QFrame#GridContainer {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
            }
        """)
        shadow_g = QGraphicsDropShadowEffect()
        shadow_g.setBlurRadius(12)
        shadow_g.setColor(QColor(0, 0, 0, 15))
        shadow_g.setOffset(0, 2)
        grid_container.setGraphicsEffect(shadow_g)

        grid_container_layout = QVBoxLayout(grid_container)
        grid_container_layout.setContentsMargins(16, 16, 16, 16)
        grid_container_layout.setSpacing(12)

        grid_title_lbl = QLabel("Transactions Ledger Mapping")
        grid_title_lbl.setStyleSheet("font-weight: 700; color: #0F172A; font-size: 14px;")
        grid_container_layout.addWidget(grid_title_lbl)

        self.grid_editor = QTableWidget()
        self.grid_editor.setAlternatingRowColors(True)
        self.grid_editor.setStyleSheet("""
            QTableWidget {
                border: none;
                gridline-color: #E2E8F0;
                font-size: 11px;
                background-color: #FFFFFF;
                alternate-background-color: #F8FAFC;
            }
            QHeaderView::section {
                background-color: #EFF6FF;
                color: #1E3A8A;
                font-weight: bold;
                border: 1px solid #DBEAFE;
                padding: 6px;
            }
        """)
        self.grid_editor.cellChanged.connect(self.on_grid_cell_changed)
        grid_container_layout.addWidget(self.grid_editor)

        main_layout.addWidget(grid_container, stretch=1)

    def close_screen(self):
        p = self.parent()
        while p:
            if hasattr(p, "switch_dashboard_page"):
                p.switch_dashboard_page("dashboard")
                break
            p = p.parent()
        self.closed.emit()

    def _safe_run_query(self, query_fn, callback_fn):
        worker = DBQueryWorker(query_fn, self)
        if not hasattr(self, "_active_workers"):
            self._active_workers = []
        self._active_workers.append(worker)
        
        def handle_result(res):
            try:
                callback_fn(res)
            finally:
                if worker in self._active_workers:
                    self._active_workers.remove(worker)
                worker.deleteLater()
                
        worker.result_ready.connect(handle_result)
        worker.start()

    def load_statements_dropdown(self):
        """Loads all completed statements for the current user into the dropdown."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else "guest"
        
        def db_query():
            col = MongoDBService.get_collection()
            if col is not None:
                return list(col.find(
                    {"user_id": user_id, "status": "Completed"},
                    sort=[("upload_date", -1)]
                ))
            return []
            
        def db_callback(logs):
            self.statement_combo.blockSignals(True)
            self.statement_combo.clear()
            self.statement_combo.addItem("Select a Statement...", "")
            
            self._statement_paths = {}
            for log in logs:
                excel_path = log.get("excel_path")
                if excel_path and os.path.exists(excel_path):
                    filename = os.path.basename(log.get("pdf_path", "Statement.pdf"))
                    upload_date = log.get("upload_date")
                    if hasattr(upload_date, "strftime"):
                        date_str = upload_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(upload_date or "")[:10]
                    display_text = f"{log.get('bank_name', 'Bank')} ({date_str}) - {filename}"
                    self.statement_combo.addItem(display_text, excel_path)
                    self._statement_paths[display_text] = excel_path
            self.statement_combo.blockSignals(False)
            
            if self.statement_combo.count() > 1 and not self.active_excel_path:
                self.statement_combo.setCurrentIndex(1)
                
        self._safe_run_query(db_query, db_callback)

    def on_statement_combo_changed(self, display_text):
        if not display_text or not hasattr(self, "_statement_paths"):
            return
        excel_path = self._statement_paths.get(display_text)
        if excel_path:
            self.on_statement_selected_by_path(excel_path)

    def on_statement_selected_by_path(self, excel_path):
        """Loads transaction data from selected Excel file and populates the table."""
        if not excel_path or not os.path.exists(excel_path):
            return
            
        self.active_excel_path = excel_path
        
        filename = os.path.basename(excel_path)
        detected_bank = "Bank Account"
        for bk in ("sbi", "hdfc", "icici", "axis", "kotak", "canara", "yes", "baroda"):
            if bk in filename.lower():
                detected_bank = f"{bk.upper()} Bank"
                break
        
        self.bank_name = detected_bank
        self.bank_ledger_input.setText(detected_bank)
        
        if hasattr(self, "_statement_paths"):
            for text, path in self._statement_paths.items():
                if path == excel_path:
                    self.statement_combo.blockSignals(True)
                    self.statement_combo.setCurrentText(text)
                    self.statement_combo.blockSignals(False)
                    break
        
        try:
            self.transactions = self.load_transactions_from_excel(excel_path)
            self.populate_grid_editor()
            Toast.success(self, "✓ Loaded statement data successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error Loading Statement", f"Could not load transaction data:\n{e}")

    def load_transactions_from_excel(self, excel_path) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        transactions = []
        
        sheet_names = wb.sheetnames
        target_sheet = None
        for name in ("Transactions Ledger", "GST Transactions", "Sheet1"):
            if name in sheet_names:
                target_sheet = wb[name]
                break
        if not target_sheet:
            target_sheet = wb.active
            
        rows = list(target_sheet.iter_rows(values_only=True))
        if not rows:
            return []
            
        header_row = rows[0]
        col_map = {}
        for idx, h in enumerate(header_row):
            if h:
                col_map[str(h).strip().lower()] = idx
                
        for row in rows[1:]:
            if not any(row):
                continue
                
            date_idx = col_map.get("date") or col_map.get("transaction date")
            narr_idx = col_map.get("narration") or col_map.get("description") or col_map.get("particulars")
            val_idx = col_map.get("value") or col_map.get("amount") or col_map.get("total amount")
            type_idx = col_map.get("type")
            debit_idx = col_map.get("debit")
            credit_idx = col_map.get("credit")
            
            tx = {}
            if date_idx is not None and date_idx < len(row):
                tx["date"] = str(row[date_idx]) if row[date_idx] is not None else ""
            if narr_idx is not None and narr_idx < len(row):
                tx["narration"] = str(row[narr_idx]) if row[narr_idx] is not None else ""
                
            debit_val = 0.0
            credit_val = 0.0
            if debit_idx is not None and debit_idx < len(row):
                try: debit_val = float(str(row[debit_idx]).replace(",", "").strip())
                except: pass
            if credit_idx is not None and credit_idx < len(row):
                try: credit_val = float(str(row[credit_idx]).replace(",", "").strip())
                except: pass
                
            if val_idx is not None and val_idx < len(row) and row[val_idx] is not None:
                try:
                    val = float(str(row[val_idx]).replace(",", "").strip())
                    if type_idx is not None and type_idx < len(row):
                        t_str = str(row[type_idx]).lower()
                        if "debit" in t_str:
                            debit_val = val
                        else:
                            credit_val = val
                    else:
                        if val < 0:
                            debit_val = abs(val)
                        else:
                            credit_val = val
                except:
                    pass
                    
            tx["debit"] = debit_val
            tx["credit"] = credit_val
            
            if type_idx is not None and type_idx < len(row) and row[type_idx]:
                tx["type"] = str(row[type_idx]).strip()
            else:
                tx["type"] = "Debit" if debit_val > 0 else "Credit"
                
            default_counter = self.counter_ledger_input.text().strip() or "Suspense Account"
            tx["voucher_type"] = TallyService.auto_derive_voucher_type(tx, tx.get("narration", ""))
            tx["ledger_name"] = default_counter
            
            transactions.append(tx)
        return transactions

    def populate_grid_editor(self):
        """Fills table widget grid using parsed transactions."""
        self._is_updating_table = True
        
        headers = ["Date", "Narration / Description", "Debit (Dr)", "Credit (Cr)", "Voucher Type", "Counter Ledger"]
        self.grid_editor.clear()
        self.grid_editor.setRowCount(len(self.transactions))
        self.grid_editor.setColumnCount(len(headers))
        self.grid_editor.setHorizontalHeaderLabels(headers)
        
        for r_idx, tx in enumerate(self.transactions):
            item_date = QTableWidgetItem(tx.get("date", ""))
            item_date.setFlags(item_date.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.grid_editor.setItem(r_idx, 0, item_date)

            item_narr = QTableWidgetItem(tx.get("narration", ""))
            item_narr.setFlags(item_narr.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.grid_editor.setItem(r_idx, 1, item_narr)

            item_deb = QTableWidgetItem(f"₹ {tx.get('debit', 0.0):,.2f}")
            item_deb.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            item_deb.setFlags(item_deb.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.grid_editor.setItem(r_idx, 2, item_deb)

            item_cred = QTableWidgetItem(f"₹ {tx.get('credit', 0.0):,.2f}")
            item_cred.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            item_cred.setFlags(item_cred.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.grid_editor.setItem(r_idx, 3, item_cred)

            vch_combo = QComboBox()
            vch_combo.addItems(["Payment", "Receipt", "Contra", "Journal"])
            vch_combo.setCurrentText(tx.get("voucher_type", "Payment"))
            vch_combo.currentTextChanged.connect(lambda text, r=r_idx: self.on_row_voucher_changed(r, text))
            self.grid_editor.setCellWidget(r_idx, 4, vch_combo)

            item_ledger = QTableWidgetItem(tx.get("ledger_name", "Suspense Account"))
            self.grid_editor.setItem(r_idx, 5, item_ledger)

        self.grid_editor.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.grid_editor.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._is_updating_table = False

    def on_row_voucher_changed(self, row, text):
        if row < 0 or row >= len(self.transactions):
            return
        self.transactions[row]["voucher_type"] = text

    def on_grid_cell_changed(self, row, col):
        if self._is_updating_table or not self.transactions:
            return
        if row < 0 or row >= len(self.transactions):
            return
            
        item = self.grid_editor.item(row, col)
        if not item:
            return
            
        new_val = item.text().strip()
        if col == 5:
            self.transactions[row]["ledger_name"] = new_val

    def on_global_ledger_changed(self):
        """Fires when user updates default bank or counter ledger inputs. Refreshes table columns."""
        if not self.transactions:
            return
            
        self.bank_name = self.bank_ledger_input.text().strip() or "Bank Account"
        default_counter = self.counter_ledger_input.text().strip() or "Suspense Account"
        
        self._is_updating_table = True
        for r_idx, tx in enumerate(self.transactions):
            old_val = tx.get("ledger_name")
            if not old_val or old_val == "Suspense Account" or old_val == self.counter_ledger_input.placeholderText():
                tx["ledger_name"] = default_counter
                item = self.grid_editor.item(r_idx, 5)
                if item:
                    item.setText(default_counter)
        self._is_updating_table = False

    def get_current_grid_transactions(self):
        """Collects the latest state from table rows and widgets."""
        current_txs = []
        for r_idx in range(len(self.transactions)):
            tx = self.transactions[r_idx].copy()
            combo = self.grid_editor.cellWidget(r_idx, 4)
            if combo:
                tx["voucher_type"] = combo.currentText()
            ledger_item = self.grid_editor.item(r_idx, 5)
            if ledger_item:
                tx["ledger_name"] = ledger_item.text().strip()
            current_txs.append(tx)
        return current_txs

    def sync_direct_to_tally_action(self):
        """
        Executes the two-stage Tally HTTP sync process:
        Stage 1: Extracts and creates ledger master accounts in Tally.
        Stage 2: Pushes accounting voucher entries.
        """
        if not self.transactions:
            QMessageBox.warning(self, "No Data", "Please select a statement first.")
            return

        server_url = self.server_url_input.text().strip()
        if not server_url:
            QMessageBox.warning(self, "Missing URL", "Please enter a valid Tally Server URL.")
            return

        bank_ledger = self.bank_ledger_input.text().strip() or "Bank Account"
        default_counter = self.counter_ledger_input.text().strip() or "Suspense Account"

        # Gathers latest grid states
        current_txs = self.get_current_grid_transactions()

        # STAGE 1: Extract all unique ledger names and map to parent groups
        ledgers_to_create = []
        
        # Include bank ledger
        ledgers_to_create.append({"name": bank_ledger, "parent": "Bank Accounts"})
        
        # Gather counter ledgers
        unique_counters = set()
        for tx in current_txs:
            ledger_name = tx.get("ledger_name") or default_counter
            unique_counters.add(ledger_name)
            
        # Determine parent groups for each counter ledger based on typical transaction types
        for ledger_name in unique_counters:
            if ledger_name.lower() == "cash":
                continue # Default Cash exists in Tally, skip
                
            # Find a matching transaction to guess group (expenses vs income)
            parent_group = "Suspense Accounts"
            for tx in current_txs:
                if (tx.get("ledger_name") or default_counter) == ledger_name:
                    vch = tx.get("voucher_type", "Payment")
                    if vch == "Payment":
                        parent_group = "Indirect Expenses"
                    elif vch == "Receipt":
                        parent_group = "Indirect Incomes"
                    elif vch == "Contra":
                        parent_group = "Cash-in-hand"
                    break
            ledgers_to_create.append({"name": ledger_name, "parent": parent_group})

        # Generate Ledger XML
        ledgers_xml = TallyService.generate_ledger_masters_xml(ledgers_to_create)
        
        self.setEnabled(False)
        Toast.info(self, "Sync Stage 1: Creating Ledgers in Tally...")

        # Push Ledgers
        ledger_result = TallyService.push_to_tally_server(server_url, ledgers_xml)
        
        if not ledger_result["success"] and "Connection to Tally failed" in ledger_result["message"]:
            self.setEnabled(True)
            QMessageBox.critical(
                self, "Sync Connection Failed", 
                f"{ledger_result['message']}\n\n"
                "Verify Tally is acts as a server on port 9000 (F1: Help -> Settings -> Connectivity)."
            )
            return

        # STAGE 2: Push Vouchers XML
        Toast.info(self, "Sync Stage 2: Syncing Vouchers to Tally...")
        vouchers_xml = TallyService.generate_tally_xml(current_txs, bank_ledger, default_counter)
        voucher_result = TallyService.push_to_tally_server(server_url, vouchers_xml)

        self.setEnabled(True)

        # Sync Outcome Summary
        if voucher_result["success"]:
            Toast.success(self, "✓ Sync with Tally completed successfully!")
            summary_msg = (
                f"Ledgers processed: {len(ledgers_to_create)}\n"
                f"Vouchers imported: {voucher_result['created']}\n\n"
                "All transaction details have been registered into Tally's books!"
            )
            QMessageBox.information(self, "Sync Success", summary_msg)
        else:
            err_msg = voucher_result["message"]
            raw_err = voucher_result["raw_response"]
            summary_msg = (
                f"Tally Voucher Sync Failed:\n\n{err_msg}\n\n"
                "Please check the log and verify your company is open in Tally.\n"
                "You can still export the XML manually using the 'Export XML' button."
            )
            QMessageBox.warning(self, "Sync Error", summary_msg)

    def export_tally_xml_action(self):
        if not self.transactions:
            QMessageBox.warning(self, "No Data", "Please select a statement first.")
            return

        current_txs = self.get_current_grid_transactions()

        default_name = f"Tally_Vouchers_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Tally XML Vouchers", default_name, "XML Files (*.xml)"
        )
        if not file_path:
            return

        try:
            bank_ledger = self.bank_ledger_input.text().strip() or "Bank Account"
            default_counter = self.counter_ledger_input.text().strip() or "Suspense Account"
            
            xml_content = TallyService.generate_tally_xml(current_txs, bank_ledger, default_counter)
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(xml_content)
                
            Toast.success(self, "✓ Tally XML vouchers exported successfully!")
            QMessageBox.information(
                self, "Export Success", 
                f"XML file generated at:\n{file_path}\n\nYou can import this in Tally via:\nImport Data -> Vouchers"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not export Tally XML vouchers:\n{e}")

    def export_tally_excel_action(self):
        if not self.transactions:
            QMessageBox.warning(self, "No Data", "Please select a statement first.")
            return

        current_txs = self.get_current_grid_transactions()

        default_name = f"Tally_Excel_Import_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Tally Excel Template", default_name, "Excel Workbooks (*.xlsx)"
        )
        if not file_path:
            return

        try:
            bank_ledger = self.bank_ledger_input.text().strip() or "Bank Account"
            default_counter = self.counter_ledger_input.text().strip() or "Suspense Account"
            
            TallyService.generate_tally_excel(current_txs, bank_ledger, default_counter, file_path)
            Toast.success(self, "✓ Tally Excel template exported successfully!")
            QMessageBox.information(
                self, "Export Success", 
                f"Excel template generated at:\n{file_path}\n\nYou can use this in Tally Prime via the Excel voucher import utility."
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not export Tally Excel sheet:\n{e}")

    def update_theme_style(self, theme):
        """Updates colors and themes dynamically to match app settings."""
        self.current_theme = theme
        if theme == "dark":
            self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #F8FAFC;")
            self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #94A3B8;")
            self.findChild(QFrame, "ControlCard").setStyleSheet("""
                QFrame#ControlCard {
                    background-color: #1E293B;
                    border: 1px solid #334155;
                    border-radius: 12px;
                }
            """)
            self.findChild(QFrame, "GridContainer").setStyleSheet("""
                QFrame#GridContainer {
                    background-color: #1E293B;
                    border: 1px solid #334155;
                    border-radius: 12px;
                }
            """)
            self.grid_editor.setStyleSheet("""
                QTableWidget {
                    border: none;
                    gridline-color: #334155;
                    font-size: 11px;
                    background-color: #1E293B;
                    alternate-background-color: #0F172A;
                    color: #F8FAFC;
                }
                QHeaderView::section {
                    background-color: #1E3A8A;
                    color: #93C5FD;
                    font-weight: bold;
                    border: 1px solid #1E293B;
                    padding: 6px;
                }
            """)
            for le in (self.bank_ledger_input, self.counter_ledger_input, self.server_url_input):
                le.setStyleSheet("""
                    QLineEdit {
                        border: 1px solid #475569;
                        border-radius: 6px;
                        padding: 6px 10px;
                        color: #F8FAFC;
                        background-color: #334155;
                    }
                """)
            self.statement_combo.setStyleSheet("""
                QComboBox {
                    border: 1px solid #475569;
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-weight: 600;
                    color: #F8FAFC;
                    background-color: #334155;
                }
            """)
        else:
            self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #0F172A;")
            self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #64748B;")
            self.findChild(QFrame, "ControlCard").setStyleSheet("""
                QFrame#ControlCard {
                    background-color: #FFFFFF;
                    border: 1px solid #E2E8F0;
                    border-radius: 12px;
                }
            """)
            self.findChild(QFrame, "GridContainer").setStyleSheet("""
                QFrame#GridContainer {
                    background-color: #FFFFFF;
                    border: 1px solid #E2E8F0;
                    border-radius: 12px;
                }
            """)
            self.grid_editor.setStyleSheet("""
                QTableWidget {
                    border: none;
                    gridline-color: #E2E8F0;
                    font-size: 11px;
                    background-color: #FFFFFF;
                    alternate-background-color: #F8FAFC;
                }
                QHeaderView::section {
                    background-color: #EFF6FF;
                    color: #1E3A8A;
                    font-weight: bold;
                    border: 1px solid #DBEAFE;
                    padding: 6px;
                }
            """)
            for le in (self.bank_ledger_input, self.counter_ledger_input, self.server_url_input):
                le.setStyleSheet("""
                    QLineEdit {
                        border: 1px solid #CBD5E1;
                        border-radius: 6px;
                        padding: 6px 10px;
                        color: #1E293B;
                        background-color: #FFFFFF;
                    }
                """)
            self.statement_combo.setStyleSheet("""
                QComboBox {
                    border: 1px solid #CBD5E1;
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-weight: 600;
                    color: #1E293B;
                    background-color: #F8FAFC;
                }
            """)
