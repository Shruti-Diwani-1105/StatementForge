import os
import json
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton,
    QComboBox, QCheckBox, QLineEdit, QScrollArea, QMessageBox, QFileDialog,
    QGraphicsDropShadowEffect, QTableWidget, QTableWidgetItem, QHeaderView,
    QRadioButton, QButtonGroup, QProgressBar
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QCursor, QPixmap, QColor, QFont

from services.duplicate_finder_service import DuplicateFinderService
from services.history_service import HistoryService
from services.excel_generator import ExcelGenerator
from utils.user_session import UserSession
from widgets.custom_button import PrimaryButton, SecondaryButton
from settings.toast import Toast


class DuplicateFinderWidget(QWidget):
    """
    Dedicated, Big-4 style Duplicate Finder & Audit Hub.
    Allows users to load single or multi-statement history logs or upload new PDFs,
    run multi-criteria duplicate detection, visually audit clusters, and export reports.
    """
    closed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_theme = "light"
        
        # State
        self.loaded_statements = [] # list of payload dicts
        self.analysis_result = None
        self.user_decisions = {}    # item_id -> 'keep' or 'remove'
        
        self.init_ui()

    def init_ui(self):
        # Set global modern QSS styles for widgets inside DuplicateFinderWidget
        self.setStyleSheet("""
            QComboBox {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding: 4px 12px;
                font-size: 13px;
                color: #0F172A;
                font-weight: 500;
            }
            QComboBox:hover {
                border-color: #2563EB;
                background-color: #FFFFFF;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QLineEdit {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding: 6px 14px;
                font-size: 13px;
                color: #0F172A;
            }
            QLineEdit:focus {
                border-color: #2563EB;
                background-color: #FFFFFF;
            }
            QCheckBox {
                font-size: 13px;
                font-weight: 600;
                color: #1E293B;
                spacing: 8px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 4px;
                border: 1.5px solid #94A3B8;
                background-color: #FFFFFF;
            }
            QCheckBox::indicator:checked {
                background-color: #2563EB;
                border-color: #2563EB;
            }
        """)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(32, 24, 32, 32)
        main_layout.setSpacing(20)

        # ==========================================
        # 1. TOP BAR: TITLE & DASHBOARD BACK BUTTON
        # ==========================================
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)
        self.title_lbl = QLabel("Duplicate Transaction Finder & Audit Hub")
        self.title_lbl.setStyleSheet("font-size: 24px; font-weight: 700; color: #0F172A;")
        self.subtitle_lbl = QLabel("Detect double-billing, duplicate uploads, and recurring entry anomalies across your bank statements.")
        self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #64748B;")
        
        text_layout.addWidget(self.title_lbl)
        text_layout.addWidget(self.subtitle_lbl)
        header_layout.addLayout(text_layout)
        header_layout.addStretch()
        
        self.close_btn = QPushButton("Back to Dashboard")
        self.close_btn.setFixedWidth(150)
        self.close_btn.setFixedHeight(34)
        self.close_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.close_btn.setStyleSheet("""
            QPushButton {
                background-color: #FFFFFF;
                color: #334155;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 0px 14px;
                font-size: 12px;
                font-weight: 600;
                min-height: 34px;
                max-height: 34px;
            }
            QPushButton:hover {
                background-color: #F8FAFC;
                border-color: #94A3B8;
            }
        """)
        self.close_btn.clicked.connect(self.close_view)
        header_layout.addWidget(self.close_btn)
        
        main_layout.addWidget(header_widget)

        # ==========================================
        # 2. STATEMENT SOURCE & SCAN CONFIG CARD
        # ==========================================
        config_card = QFrame()
        config_card.setObjectName("ConfigCard")
        config_card.setStyleSheet("""
            QFrame#ConfigCard {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 14px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(16)
        shadow.setColor(QColor(0, 0, 0, 14))
        shadow.setOffset(0, 3)
        config_card.setGraphicsEffect(shadow)

        cfg_layout = QVBoxLayout(config_card)
        cfg_layout.setContentsMargins(22, 18, 22, 18)
        cfg_layout.setSpacing(16)

        # Source Selection Row
        src_row = QHBoxLayout()
        src_row.setSpacing(12)

        src_label = QLabel("Statement Source:")
        src_label.setStyleSheet("font-weight: 700; color: #0F172A; font-size: 13px;")
        src_row.addWidget(src_label)

        self.history_combo = QComboBox()
        self.history_combo.setMinimumWidth(300)
        self.history_combo.setFixedHeight(34)
        self.history_combo.setStyleSheet("""
            QComboBox {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 0px 12px;
                font-size: 12px;
                color: #0F172A;
                min-height: 34px;
                max-height: 34px;
            }
            QComboBox:hover {
                border-color: #2563EB;
                background-color: #FFFFFF;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
        """)
        self.history_combo.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        src_row.addWidget(self.history_combo)

        btn_style_secondary = """
            QPushButton {
                background-color: #FFFFFF;
                color: #334155;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 0px 14px;
                font-size: 12px;
                font-weight: 600;
                min-height: 34px;
                max-height: 34px;
            }
            QPushButton:hover {
                background-color: #F8FAFC;
                border-color: #94A3B8;
            }
        """

        btn_style_primary = """
            QPushButton {
                background-color: #2563EB;
                color: #FFFFFF;
                border: none;
                border-radius: 6px;
                padding: 0px 14px;
                font-size: 12px;
                font-weight: 600;
                min-height: 34px;
                max-height: 34px;
            }
            QPushButton:hover {
                background-color: #1D4ED8;
            }
        """

        self.btn_load_history = QPushButton("Load Selected")
        self.btn_load_history.setFixedHeight(34)
        self.btn_load_history.setStyleSheet(btn_style_secondary)
        self.btn_load_history.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_load_history.clicked.connect(self.load_from_history)
        src_row.addWidget(self.btn_load_history)

        self.btn_multi_audit = QPushButton("Cross-Statement Scan")
        self.btn_multi_audit.setFixedHeight(34)
        self.btn_multi_audit.setStyleSheet(btn_style_secondary)
        self.btn_multi_audit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_multi_audit.clicked.connect(self.load_all_history_statements)
        src_row.addWidget(self.btn_multi_audit)

        self.btn_upload = QPushButton("Upload New PDF")
        self.btn_upload.setFixedHeight(34)
        self.btn_upload.setStyleSheet(btn_style_primary)
        self.btn_upload.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_upload.clicked.connect(self.upload_new_statement)
        src_row.addWidget(self.btn_upload)

        src_row.addStretch()
        cfg_layout.addLayout(src_row)

        # Divider line
        div_line = QFrame()
        div_line.setFrameShape(QFrame.Shape.HLine)
        div_line.setStyleSheet("color: #F1F5F9;")
        cfg_layout.addWidget(div_line)

        # Rules Options Row
        rules_row = QHBoxLayout()
        rules_row.setSpacing(20)

        rules_label = QLabel("Scan Criteria:")
        rules_label.setStyleSheet("font-weight: 700; color: #0F172A; font-size: 13px;")
        rules_row.addWidget(rules_label)

        self.chk_exact = QCheckBox("Exact Match (100%)")
        self.chk_exact.setChecked(True)
        rules_row.addWidget(self.chk_exact)

        self.chk_potential = QCheckBox("Potential / Fuzzy Match")
        self.chk_potential.setChecked(True)
        rules_row.addWidget(self.chk_potential)

        # Date Window
        date_win_lbl = QLabel("Date Window:")
        date_win_lbl.setStyleSheet("color: #475569; font-size: 12px; font-weight: 600;")
        rules_row.addWidget(date_win_lbl)

        self.combo_date_win = QComboBox()
        self.combo_date_win.addItems(["Same Date (0 days)", "± 1 Day", "± 2 Days", "± 3 Days"])
        self.combo_date_win.setCurrentIndex(2) # Default ±2 days
        self.combo_date_win.setFixedWidth(145)
        self.combo_date_win.setFixedHeight(34)
        self.combo_date_win.setStyleSheet("""
            QComboBox {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 0px 10px;
                font-size: 12px;
                color: #0F172A;
                min-height: 34px;
                max-height: 34px;
            }
        """)
        rules_row.addWidget(self.combo_date_win)

        # Similarity Threshold
        sim_lbl = QLabel("Similarity:")
        sim_lbl.setStyleSheet("color: #475569; font-size: 12px; font-weight: 600;")
        rules_row.addWidget(sim_lbl)

        self.combo_sim = QComboBox()
        self.combo_sim.addItems(["Strict (85%)", "Standard (75%)", "Flexible (60%)"])
        self.combo_sim.setCurrentIndex(1)
        self.combo_sim.setFixedWidth(135)
        self.combo_sim.setFixedHeight(34)
        self.combo_sim.setStyleSheet("""
            QComboBox {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 0px 10px;
                font-size: 12px;
                color: #0F172A;
                min-height: 34px;
                max-height: 34px;
            }
        """)
        rules_row.addWidget(self.combo_sim)

        self.btn_run_scan = QPushButton("Run Scan")
        self.btn_run_scan.setFixedWidth(110)
        self.btn_run_scan.setFixedHeight(34)
        self.btn_run_scan.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_run_scan.setStyleSheet(btn_style_primary)
        self.btn_run_scan.clicked.connect(self.run_duplicate_scan)
        rules_row.addWidget(self.btn_run_scan)

        rules_row.addStretch()
        cfg_layout.addLayout(rules_row)

        main_layout.addWidget(config_card)

        # ==========================================
        # 3. KPI SUMMARY METRIC CARDS ROW
        # ==========================================
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(16)

        self.card_total_tx = self._create_kpi_card("Transactions Scanned", "0", "Total items in statement", "#EFF6FF", "#2563EB")
        self.card_dup_clusters = self._create_kpi_card("Duplicate Clusters", "0", "Identified anomaly sets", "#FEF2F2", "#EF4444")
        self.card_flagged_amt = self._create_kpi_card("Flagged Monetary Amount", "₹ 0.00", "Potential double billing", "#FFFBEB", "#D97706")
        self.card_clean_score = self._create_kpi_card("Cleanliness Score", "100%", "Statement data integrity", "#F0FDF4", "#16A34A")

        kpi_row.addWidget(self.card_total_tx)
        kpi_row.addWidget(self.card_dup_clusters)
        kpi_row.addWidget(self.card_flagged_amt)
        kpi_row.addWidget(self.card_clean_score)

        main_layout.addLayout(kpi_row)

        # ==========================================
        # 4. FILTER & AUTO-RESOLVE TOOLBAR
        # ==========================================
        filter_card = QFrame()
        filter_card.setStyleSheet("background-color: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 10px; padding: 4px;")
        filter_lay = QHBoxLayout(filter_card)
        filter_lay.setContentsMargins(14, 8, 14, 8)
        filter_lay.setSpacing(12)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Search clusters by narration, amount, ref no...")
        self.search_input.setFixedHeight(36)
        self.search_input.textChanged.connect(self.render_clusters)
        filter_lay.addWidget(self.search_input)

        self.filter_type_combo = QComboBox()
        self.filter_type_combo.addItems(["All Match Types", "Exact Match Only", "Potential Match Only", "Cross-Statement Only"])
        self.filter_type_combo.setFixedWidth(180)
        self.filter_type_combo.setFixedHeight(36)
        self.filter_type_combo.currentIndexChanged.connect(self.render_clusters)
        filter_lay.addWidget(self.filter_type_combo)

        filter_lay.addStretch()

        # Preset Auto-Resolution Action Buttons
        btn_resolve_first = SecondaryButton("Keep First Entry")
        btn_resolve_first.setFixedHeight(36)
        btn_resolve_first.clicked.connect(lambda: self.apply_preset_resolution("keep_first"))
        filter_lay.addWidget(btn_resolve_first)

        btn_resolve_last = SecondaryButton("Keep Last Entry")
        btn_resolve_last.setFixedHeight(36)
        btn_resolve_last.clicked.connect(lambda: self.apply_preset_resolution("keep_last"))
        filter_lay.addWidget(btn_resolve_last)

        main_layout.addWidget(filter_card)

        # ==========================================
        # 5. SCROLLABLE DUPLICATE CLUSTERS AREA
        # ==========================================
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll_area.setStyleSheet("background: transparent;")

        self.scroll_container = QWidget()
        self.scroll_container.setStyleSheet("background: transparent;")
        self.clusters_layout = QVBoxLayout(self.scroll_container)
        self.clusters_layout.setContentsMargins(0, 0, 0, 0)
        self.clusters_layout.setSpacing(16)

        self.scroll_area.setWidget(self.scroll_container)
        main_layout.addWidget(self.scroll_area, stretch=1)

        # ==========================================
        # 6. BOTTOM ACTION & EXPORT BAR
        # ==========================================
        bottom_bar = QFrame()
        bottom_bar.setStyleSheet("background-color: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 12px; padding: 12px;")
        bot_layout = QHBoxLayout(bottom_bar)
        bot_layout.setContentsMargins(20, 10, 20, 10)
        bot_layout.setSpacing(16)

        self.resolution_status_lbl = QLabel("No audit scan performed yet.")
        self.resolution_status_lbl.setStyleSheet("""
            background-color: #F1F5F9;
            color: #1E293B;
            font-weight: 600;
            font-size: 13px;
            padding: 6px 14px;
            border-radius: 20px;
            border: 1px solid #E2E8F0;
        """)
        bot_layout.addWidget(self.resolution_status_lbl)

        bot_layout.addStretch()

        self.btn_export_audit = QPushButton("Export Audit Report (.xlsx)")
        self.btn_export_audit.setFixedWidth(220)
        self.btn_export_audit.setFixedHeight(40)
        self.btn_export_audit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export_audit.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: #FFFFFF;
                font-weight: 700;
                font-size: 13px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background-color: #047857;
            }
        """)
        self.btn_export_audit.clicked.connect(self.export_audit_excel)
        bot_layout.addWidget(self.btn_export_audit)

        self.btn_export_cleaned = QPushButton("Export Cleaned Excel (.xlsx)")
        self.btn_export_cleaned.setFixedWidth(230)
        self.btn_export_cleaned.setFixedHeight(40)
        self.btn_export_cleaned.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_export_cleaned.setStyleSheet("""
            QPushButton {
                background-color: #2563EB;
                color: #FFFFFF;
                font-weight: 700;
                font-size: 13px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background-color: #1D4ED8;
            }
        """)
        self.btn_export_cleaned.clicked.connect(self.export_cleaned_excel)
        bot_layout.addWidget(self.btn_export_cleaned)

        self.btn_send_email = QPushButton("✉ Send via Email")
        self.btn_send_email.setFixedWidth(170)
        self.btn_send_email.setFixedHeight(40)
        self.btn_send_email.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_send_email.setStyleSheet("""
            QPushButton {
                background-color: #7C3AED;
                color: #FFFFFF;
                font-weight: 700;
                font-size: 13px;
                border-radius: 8px;
                border: none;
            }
            QPushButton:hover {
                background-color: #6D28D9;
            }
        """)
        self.btn_send_email.clicked.connect(self.open_email_composer)
        bot_layout.addWidget(self.btn_send_email)

        main_layout.addWidget(bottom_bar)

        # Initial load of history dropdown
        self.load_history_dropdown()

    # --- UI Helpers ---

    def _create_kpi_card(self, title, value, subtitle, bg_color, accent_color):
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 12))
        shadow.setOffset(0, 3)
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(6)

        # Top Accent Line
        accent = QFrame()
        accent.setFixedHeight(4)
        accent.setStyleSheet(f"background-color: {accent_color}; border-radius: 2px;")
        layout.addWidget(accent)

        t_lbl = QLabel(title)
        t_lbl.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B; letter-spacing: 0.5px; text-transform: uppercase;")
        layout.addWidget(t_lbl)

        v_lbl = QLabel(value)
        v_lbl.setObjectName("KpiValueLabel")
        v_lbl.setStyleSheet(f"font-size: 24px; font-weight: 800; color: {accent_color}; margin-top: 2px;")
        layout.addWidget(v_lbl)

        sub_lbl = QLabel(subtitle)
        sub_lbl.setStyleSheet("font-size: 11px; color: #94A3B8; font-weight: 500;")
        layout.addWidget(sub_lbl)

        return card

    def load_history_dropdown(self):
        """Populates statement history dropdown."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else None
        
        self.history_combo.clear()
        self.history_combo.addItem("Select statement from history...", None)

        logs = HistoryService.get_history_logs(user_id)
        completed_logs = [l for l in logs if l.get("status") == "Completed" and l.get("excel_path")]

        for log in completed_logs:
            clean_log = {}
            for k, v in log.items():
                if isinstance(v, (datetime.datetime, datetime.date)):
                    clean_log[k] = v.strftime("%Y-%m-%d")
                else:
                    clean_log[k] = str(v) if not isinstance(v, (str, int, float, bool, list, dict, type(None))) else v

            bank = clean_log.get("bank_name", "Unknown Bank")
            date_str = str(clean_log.get("upload_date") or "")[:10]
            tx_count = clean_log.get("total_transactions", 0)
            excel_path = clean_log.get("excel_path", "")
            disp = f"{bank} ({date_str}) - {tx_count} txs"
            self.history_combo.addItem(disp, clean_log)

    def load_from_history(self):
        """Loads transaction data from selected history item excel file."""
        log = self.history_combo.currentData()
        if not log:
            Toast.display_toast(self, "Please select a valid statement from history.", toast_type="warning")
            return

        excel_path = log.get("excel_path", "")
        if not os.path.exists(excel_path):
            Toast.display_toast(self, "Associated statement file could not be found.", toast_type="error")
            return

        transactions = self._read_transactions_from_excel(excel_path)
        if not transactions:
            Toast.display_toast(self, "Could not extract transactions from statement Excel file.", toast_type="error")
            return

        payload = {
            "file_name": os.path.basename(excel_path),
            "bank_name": log.get("bank_name", "Bank"),
            "transactions": transactions
        }
        self.loaded_statements = [payload]
        Toast.display_toast(self, f"Loaded {len(transactions)} transactions from history.", toast_type="success")
        self.run_duplicate_scan()

    def load_all_history_statements(self):
        """Loads transactions from ALL completed history logs for multi-statement cross audit."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else None
        logs = HistoryService.get_history_logs(user_id)
        completed_logs = [l for l in logs if l.get("status") == "Completed" and l.get("excel_path")]

        if not completed_logs:
            Toast.display_toast(self, "No completed statements available in history.", toast_type="warning")
            return

        payloads = []
        for log in completed_logs:
            path = log.get("excel_path", "")
            if os.path.exists(path):
                txs = self._read_transactions_from_excel(path)
                if txs:
                    payloads.append({
                        "file_name": os.path.basename(path),
                        "bank_name": log.get("bank_name", "Bank"),
                        "transactions": txs
                    })

        if not payloads:
            Toast.display_toast(self, "Failed to read history statement files.", toast_type="error")
            return

        self.loaded_statements = payloads
        total_tx = sum(len(p["transactions"]) for p in payloads)
        Toast.display_toast(self, f"Loaded {len(payloads)} statements ({total_tx} total transactions) for Cross Audit.", toast_type="success")
        self.run_duplicate_scan()

    def upload_new_statement(self):
        """Allows user to upload and parse a PDF directly into Duplicate Finder."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Bank Statement PDF", "", "PDF Files (*.pdf)")
        if not file_path:
            return

        from parser.parser import PDFStatementParser
        try:
            payload = PDFStatementParser.parse(file_path)
            txs = payload.get("transactions", [])
            if not txs:
                Toast.display_toast(self, "No transactions extracted from uploaded PDF.", toast_type="warning")
                return

            self.loaded_statements = [{
                "file_name": os.path.basename(file_path),
                "bank_name": payload.get("bank_name", "Bank"),
                "transactions": txs
            }]
            Toast.display_toast(self, f"Parsed {len(txs)} transactions from PDF.", toast_type="success")
            self.run_duplicate_scan()
        except Exception as e:
            QMessageBox.critical(self, "Parsing Error", f"Failed to parse uploaded PDF statement:\n{e}")

    def _read_transactions_from_excel(self, excel_path):
        """Reads transactions back from an openpyxl Excel file."""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
            txs = []
            header_found = False
            col_map = {}

            for row in sheet.iter_rows(values_only=True):
                if not row:
                    continue
                row_str = [str(c).lower().strip() if c is not None else "" for c in row]
                
                # Check for header row
                if not header_found and any(h in name for name in row_str for h in ["date", "narration", "description"]):
                    header_found = True
                    for i, name in enumerate(row_str):
                        if "date" in name and "value" not in name: col_map["date"] = i
                        elif "narration" in name or "particular" in name or "description" in name: col_map["narration"] = i
                        elif "debit" in name or "withdrawal" in name: col_map["debit"] = i
                        elif "credit" in name or "deposit" in name: col_map["credit"] = i
                        elif "balance" in name: col_map["balance"] = i
                        elif "ref" in name or "cheque" in name or "chq" in name: col_map["ref_no"] = i
                    continue

                if header_found and "date" in col_map:
                    date_val = str(row[col_map["date"]]).strip() if len(row) > col_map["date"] and row[col_map["date"]] is not None else ""
                    if not date_val or date_val.lower() in ["date", "none", "null", ""] or "total" in date_val.lower():
                        continue

                    narr_val = str(row[col_map["narration"]]).strip() if "narration" in col_map and len(row) > col_map["narration"] and row[col_map["narration"]] is not None else ""
                    deb_val = str(row[col_map["debit"]]).strip() if "debit" in col_map and len(row) > col_map["debit"] and row[col_map["debit"]] is not None else ""
                    cred_val = str(row[col_map["credit"]]).strip() if "credit" in col_map and len(row) > col_map["credit"] and row[col_map["credit"]] is not None else ""
                    bal_val = str(row[col_map["balance"]]).strip() if "balance" in col_map and len(row) > col_map["balance"] and row[col_map["balance"]] is not None else ""
                    ref_val = str(row[col_map["ref_no"]]).strip() if "ref_no" in col_map and len(row) > col_map["ref_no"] and row[col_map["ref_no"]] is not None else ""

                    txs.append({
                        "date": date_val,
                        "narration": narr_val,
                        "debit": deb_val if deb_val.lower() != "none" else "",
                        "credit": cred_val if cred_val.lower() != "none" else "",
                        "balance": bal_val if bal_val.lower() != "none" else "",
                        "ref_no": ref_val if ref_val.lower() != "none" else ""
                    })

            return txs
        except Exception as e:
            print(f"Error reading excel: {e}")
            return []

    # --- Scanning & Cluster Rendering ---

    def run_duplicate_scan(self):
        """Runs the DuplicateFinderService analysis with current UI rules."""
        if not self.loaded_statements:
            Toast.display_toast(self, "Please select or upload a statement first.", toast_type="warning")
            return

        date_win_map = [0, 1, 2, 3]
        sim_map = [0.85, 0.75, 0.60]

        options = {
            "exact_match": self.chk_exact.isChecked(),
            "potential_match": self.chk_potential.isChecked(),
            "date_window_days": date_win_map[self.combo_date_win.currentIndex()],
            "similarity_threshold": sim_map[self.combo_sim.currentIndex()]
        }

        if len(self.loaded_statements) == 1:
            self.analysis_result = DuplicateFinderService.analyze_statement(
                self.loaded_statements[0]["transactions"],
                options
            )
        else:
            self.analysis_result = DuplicateFinderService.analyze_multiple_statements(
                self.loaded_statements,
                options
            )

        # Default resolution decision: Keep first entry
        clusters = self.analysis_result.get("clusters", [])
        self.user_decisions = DuplicateFinderService.apply_auto_resolution(clusters, strategy="keep_first")

        self.update_kpi_cards()
        self.render_clusters()

    def update_kpi_cards(self):
        if not self.analysis_result:
            return

        stats = self.analysis_result.get("stats", {})
        
        # Total Tx
        val_total = self.card_total_tx.findChild(QLabel, "KpiValueLabel")
        if val_total: val_total.setText(str(stats.get("total_transactions", 0)))

        # Clusters
        val_clusters = self.card_dup_clusters.findChild(QLabel, "KpiValueLabel")
        if val_clusters: val_clusters.setText(str(stats.get("duplicate_clusters", 0)))

        # Amount
        val_amt = self.card_flagged_amt.findChild(QLabel, "KpiValueLabel")
        if val_amt: val_amt.setText(f"₹ {stats.get('flagged_debit_sum', 0.0) + stats.get('flagged_credit_sum', 0.0):,.2f}")

        # Score
        val_score = self.card_clean_score.findChild(QLabel, "KpiValueLabel")
        if val_score: val_score.setText(f"{stats.get('cleanliness_score', 100.0)}%")

        removals_count = sum(1 for v in self.user_decisions.values() if v == "remove")
        self.resolution_status_lbl.setText(f"Audit Complete: {stats.get('duplicate_clusters', 0)} duplicate clusters detected. {removals_count} entries flagged for removal.")

    def render_clusters(self):
        """Clears and re-renders cluster cards based on search query and filter selection."""
        # Clear existing layout safely
        if hasattr(self, "clusters_layout") and self.clusters_layout is not None:
            while self.clusters_layout.count():
                item = self.clusters_layout.takeAt(0)
                w = item.widget()
                if w is not None:
                    w.setParent(None)
                    w.deleteLater()

        if not self.analysis_result or not self.analysis_result.get("clusters"):
            empty_card = QFrame()
            empty_card.setStyleSheet("background-color: #FFFFFF; border: 1px dashed #CBD5E1; border-radius: 12px; padding: 40px;")
            empty_lay = QVBoxLayout(empty_card)
            empty_lbl = QLabel("✨ Clean Statement: No duplicate transaction anomalies detected.")
            empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_lbl.setStyleSheet("font-size: 15px; font-weight: 600; color: #16A34A;")
            empty_lay.addWidget(empty_lbl)
            self.clusters_layout.addWidget(empty_card)
            self.clusters_layout.addStretch()
            return

        search_query = self.search_input.text().lower().strip()
        filter_idx = self.filter_type_combo.currentIndex() # 0: All, 1: Exact, 2: Potential, 3: Cross

        rendered_count = 0
        for cluster in self.analysis_result.get("clusters", []):
            m_type = cluster["match_type"]

            # Filter Check
            if filter_idx == 1 and m_type != "Exact Match": continue
            if filter_idx == 2 and m_type != "Potential Duplicate": continue
            if filter_idx == 3 and m_type != "Cross-Statement Duplicate": continue

            # Search Check
            if search_query:
                match_found = False
                if search_query in cluster["title"].lower() or search_query in cluster["reason"].lower():
                    match_found = True
                else:
                    for it in cluster["items"]:
                        if (search_query in it["raw_narration"].lower() or
                            search_query in it["date_str"].lower() or
                            search_query in str(it["amount"]) or
                            search_query in it["ref_no"].lower()):
                            match_found = True
                            break
                if not match_found:
                    continue

            # Render Cluster Card
            cluster_card = self._build_cluster_card(cluster)
            self.clusters_layout.addWidget(cluster_card)
            rendered_count += 1

        if rendered_count == 0:
            no_match_lbl = QLabel("No duplicate clusters match your search/filter criteria.")
            no_match_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            no_match_lbl.setStyleSheet("font-size: 13px; color: #64748B; padding: 20px;")
            self.clusters_layout.addWidget(no_match_lbl)

        self.clusters_layout.addStretch()

    def _build_cluster_card(self, cluster):
        m_type = cluster.get("match_type", "")
        left_accent = "#10B981" if m_type == "Exact Match" else ("#F59E0B" if m_type == "Potential Duplicate" else "#6366F1")

        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-left: 5px solid {left_accent};
                border-radius: 12px;
                margin-bottom: 10px;
            }}
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setColor(QColor(0, 0, 0, 10))
        shadow.setOffset(0, 2)
        card.setGraphicsEffect(shadow)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(12)

        # Header Row
        head_lay = QHBoxLayout()
        head_lay.setSpacing(10)

        badge = QLabel(m_type.upper())
        badge.setStyleSheet(f"""
            background-color: {cluster['badge_bg']};
            color: {cluster['badge_color']};
            font-weight: 800;
            font-size: 10px;
            letter-spacing: 0.5px;
            padding: 4px 12px;
            border-radius: 12px;
        """)
        head_lay.addWidget(badge)

        title_lbl = QLabel(f"{cluster['title']}  •  {cluster['confidence']}% Confidence")
        title_lbl.setStyleSheet("font-weight: 700; font-size: 15px; color: #0F172A;")
        head_lay.addWidget(title_lbl)

        head_lay.addStretch()

        if cluster.get("reason"):
            reason_lbl = QLabel(cluster.get("reason"))
            reason_lbl.setStyleSheet("font-size: 12px; color: #475569; background-color: #F1F5F9; padding: 4px 10px; border-radius: 6px;")
            head_lay.addWidget(reason_lbl)

        items_cnt_lbl = QLabel(f"{len(cluster['items'])} Entries")
        items_cnt_lbl.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B; background-color: #F8FAFC; border: 1px solid #E2E8F0; padding: 4px 8px; border-radius: 6px;")
        head_lay.addWidget(items_cnt_lbl)

        layout.addLayout(head_lay)

        # Items Layout Container
        items_frame = QFrame()
        items_frame.setStyleSheet("background-color: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 10px;")
        items_lay = QVBoxLayout(items_frame)
        items_lay.setContentsMargins(10, 8, 10, 8)
        items_lay.setSpacing(6)

        # Table Header Row (Dark Slate Bar)
        hdr_row = QFrame()
        hdr_row.setStyleSheet("background-color: #1E293B; border-radius: 6px;")
        hdr_lay = QHBoxLayout(hdr_row)
        hdr_lay.setContentsMargins(12, 6, 12, 6)
        hdr_lay.setSpacing(12)

        def make_hdr(txt, width=None):
            lbl = QLabel(txt)
            lbl.setStyleSheet("font-weight: 700; font-size: 11px; color: #FFFFFF; letter-spacing: 0.5px;")
            if width: lbl.setFixedWidth(width)
            return lbl

        hdr_lay.addWidget(make_hdr("ACTION", 95))
        hdr_lay.addWidget(make_hdr("DATE", 95))
        hdr_lay.addWidget(make_hdr("SOURCE STATEMENT", 150))
        hdr_lay.addWidget(make_hdr("NARRATION / DESCRIPTION", None), stretch=1)
        hdr_lay.addWidget(make_hdr("AMOUNT (₹)", 130))
        hdr_lay.addWidget(make_hdr("REF NO", 100))
        items_lay.addWidget(hdr_row)

        for r_idx, item in enumerate(cluster["items"]):
            it_id = item["id"]
            row_frame = QFrame()
            row_frame.setStyleSheet("""
                QFrame {
                    background-color: #FFFFFF;
                    border: 1px solid #E2E8F0;
                    border-radius: 6px;
                }
                QFrame:hover {
                    border-color: #CBD5E1;
                    background-color: #FFFFFF;
                }
            """)
            r_lay = QHBoxLayout(row_frame)
            r_lay.setContentsMargins(10, 8, 10, 8)
            r_lay.setSpacing(12)

            action_combo = QComboBox()
            action_combo.addItems(["Keep", "Remove"])
            action_combo.setFixedWidth(95)
            action_combo.setFixedHeight(30)
            action_combo.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

            current_dec = self.user_decisions.get(it_id, "keep" if r_idx == 0 else "remove")
            action_combo.setCurrentIndex(0 if current_dec == "keep" else 1)
            action_combo.activated.connect(lambda idx, item_id=it_id: self._on_item_action_changed(item_id, idx))
            r_lay.addWidget(action_combo)

            lbl_date = QLabel(item.get("date_str", ""))
            lbl_date.setFixedWidth(95)
            lbl_date.setStyleSheet("font-size: 12px; font-weight: 600; color: #334155;")
            r_lay.addWidget(lbl_date)

            src_filename = item.get("source_file", "")
            lbl_src = QLabel(src_filename[:22] + "..." if len(src_filename) > 22 else src_filename)
            lbl_src.setToolTip(src_filename)
            lbl_src.setFixedWidth(150)
            lbl_src.setStyleSheet("font-size: 11px; font-weight: 600; color: #1E40AF; background-color: #EFF6FF; border: 1px solid #DBEAFE; padding: 2px 6px; border-radius: 4px;")
            r_lay.addWidget(lbl_src)

            lbl_narr = QLabel(item.get("raw_narration", ""))
            lbl_narr.setWordWrap(True)
            lbl_narr.setStyleSheet("font-size: 12px; color: #0F172A; font-weight: 600;")
            r_lay.addWidget(lbl_narr, stretch=1)

            tx_type = item.get("type", "").lower()
            amt_val = item.get('amount', 0.0)
            amt_color = "#DC2626" if "debit" in tx_type or "dr" in tx_type or amt_val < 0 else "#059669"
            amt_prefix = "-" if "debit" in tx_type or "dr" in tx_type else "+"
            amt_str = f"{amt_prefix}₹ {abs(amt_val):,.2f}"
            
            lbl_amt = QLabel(amt_str)
            lbl_amt.setFixedWidth(130)
            lbl_amt.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            lbl_amt.setStyleSheet(f"font-size: 13px; font-weight: 800; color: {amt_color}; padding-right: 8px;")
            r_lay.addWidget(lbl_amt)

            ref_str = item.get("ref_no", "-")
            lbl_ref = QLabel(ref_str if ref_str else "-")
            lbl_ref.setFixedWidth(100)
            lbl_ref.setStyleSheet("font-size: 11px; color: #64748B; background-color: #F8FAFC; padding: 2px 6px; border-radius: 4px;")
            r_lay.addWidget(lbl_ref)

            items_lay.addWidget(row_frame)

        layout.addWidget(items_frame)
        return card

    def _on_item_action_changed(self, item_id, idx):
        self.user_decisions[item_id] = "keep" if idx == 0 else "remove"
        removals_count = sum(1 for v in self.user_decisions.values() if v == "remove")
        self.resolution_status_lbl.setText(f"Audit Updated: {removals_count} duplicate entries flagged for removal.")

    def apply_preset_resolution(self, strategy):
        """Applies auto-resolution preset across all clusters."""
        if not self.analysis_result or not self.analysis_result.get("clusters"):
            return
        
        clusters = self.analysis_result.get("clusters", [])
        self.user_decisions = DuplicateFinderService.apply_auto_resolution(clusters, strategy=strategy)
        Toast.display_toast(self, f"Applied preset resolution ({strategy.replace('_', ' ').title()}).", toast_type="info")
        self.render_clusters()
        self.update_kpi_cards()

    # --- Export Actions ---

    def export_audit_excel(self):
        """Exports detailed Excel Audit Report."""
        if not self.analysis_result or not self.analysis_result.get("clusters"):
            Toast.display_toast(self, "No scan results available to export.", toast_type="warning")
            return

        out_path, _ = QFileDialog.getSaveFileName(self, "Save Duplicate Audit Report", "Duplicate_Transaction_Audit_Report.xlsx", "Excel Files (*.xlsx)")
        if not out_path:
            return

        try:
            clusters = self.analysis_result.get("clusters", [])
            stats = self.analysis_result.get("stats", {})
            DuplicateFinderService.export_duplicate_report(clusters, stats, out_path)
            Toast.display_toast(self, "Duplicate Audit Report exported successfully!", toast_type="success")
            QMessageBox.information(self, "Export Successful", f"Audit report saved to:\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export audit report:\n{e}")

    def export_cleaned_excel(self):
        """Exports clean statement Excel file with selected duplicates removed."""
        if not self.analysis_result or not self.analysis_result.get("annotated_transactions"):
            Toast.display_toast(self, "No statement data loaded for export.", toast_type="warning")
            return

        annotated = self.analysis_result.get("annotated_transactions", [])
        # Filter out items marked as remove
        cleaned_txs = []
        for item in annotated:
            if self.user_decisions.get(item["id"]) != "remove":
                cleaned_txs.append(item["raw"])

        out_path, _ = QFileDialog.getSaveFileName(self, "Save Cleaned Excel Statement", "Statement_Deduplicated.xlsx", "Excel Files (*.xlsx)")
        if not out_path:
            return

        try:
            bank_name = self.loaded_statements[0].get("bank_name", "Bank") if self.loaded_statements else "Bank"
            ExcelGenerator.generate_excel(
                pdf_path=out_path,
                bank_name=bank_name,
                account_holder="Statement Owner",
                period="Current Period",
                transactions=cleaned_txs,
                custom_output_path=out_path
            )
            Toast.display_toast(self, f"Cleaned statement saved with {len(cleaned_txs)} transactions!", toast_type="success")
            QMessageBox.information(self, "Export Successful", f"Cleaned statement ({len(cleaned_txs)} transactions) exported to:\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export cleaned statement:\n{e}")

    def open_email_composer(self):
        """Opens Email Composer pre-attaching active Duplicate Audit report."""
        from ui.email_composer_dialog import EmailComposerDialog
        
        attachment = getattr(self, "excel_path", None)
        bank = self.loaded_statements[0].get("bank_name", "") if hasattr(self, "loaded_statements") and self.loaded_statements else ""

        dialog = EmailComposerDialog(
            report_type="Duplicate Transaction Report",
            default_attachment=attachment,
            bank_name=bank,
            parent=self
        )
        dialog.exec()

    def close_view(self):
        self.closed.emit()
