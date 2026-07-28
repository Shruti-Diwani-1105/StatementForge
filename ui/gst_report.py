import os
import csv
import datetime
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, QPushButton,
    QTextBrowser, QFileDialog, QMessageBox, QGraphicsDropShadowEffect,
    QTabWidget, QTableWidget, QTableWidgetItem, QComboBox, QHeaderView
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QCursor, QTextDocument, QColor, QPageLayout, QPageSize
from PyQt6.QtCore import QMarginsF
from PyQt6.QtPrintSupport import QPrinter

from PyQt6.QtCore import QThread
from services.gst_service import GSTService
from widgets.custom_button import PrimaryButton, SecondaryButton
from settings.toast import Toast
from services.mongodb_service import MongoDBService
from utils.user_session import UserSession


class DBQueryWorker(QThread):
    result_ready = pyqtSignal(object)

    def __init__(self, query_fn, parent=None):
        super().__init__(parent)
        self.query_fn = query_fn

    def run(self):
        try:
            res = self.query_fn()
            self.result_ready.emit(res)
        except Exception as e:
            print(f"DBQueryWorker error: {e}")
            self.result_ready.emit(None)


class GSTR2BReconcileWorker(QThread):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, gst_ledger, filepath, parent=None):
        super().__init__(parent)
        self.gst_ledger = gst_ledger
        self.filepath = filepath

    def run(self):
        try:
            summary = GSTService.reconcile_with_gstr2b(self.gst_ledger, self.filepath)
            self.finished.emit(summary)
        except Exception as e:
            self.error.emit(str(e))




class GSTReportWidget(QWidget):
    """
    Renders an interactive, Big-4 style AI-Generated GST Reconciliation Report.
    Features date range filtering, GSTR-2B 3-way reconciliation, live grid editing, and exports.
    """
    closed = pyqtSignal()

    CATEGORIES_LIST = [
        "Bank Charges", "Processing Fees", "Service Charges", "Courier Charges", 
        "Office Expenses", "Utilities", "Software Subscription", "Vendor Payment", 
        "Fuel", "Travel", "Miscellaneous", "Personal"
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_theme = "light"
        
        # State Data
        self.parsed_payload = None
        self.gst_ledger = []
        self.excel_path = None
        self.gstr2b_summary = None
        self.gstr2b_path = None
        self.current_date_filter = "All Dates"
        self._is_updating_table = False
        
        self.init_ui()
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(500, self.load_statements_dropdown)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(24, 24, 24, 24)
        main_layout.setSpacing(16)

        # ==========================================
        # TOP BAR: TITLE & DESCRIPTION
        # ==========================================
        header_widget = QWidget()
        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 0)
        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)
        self.title_lbl = QLabel("GST Tax Ledger Reconciliation & 3-Way Audit")
        self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #0F172A;")
        self.subtitle_lbl = QLabel("Filter date ranges, import GSTR-2B purchase files, edit ledger items, and export reports.")
        self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #64748B;")
        
        text_layout.addWidget(self.title_lbl)
        text_layout.addWidget(self.subtitle_lbl)
        header_layout.addLayout(text_layout)
        header_layout.addStretch()
        
        # Close / Return button
        self.close_btn = SecondaryButton("Back to Dashboard")
        self.close_btn.setFixedWidth(150)
        self.close_btn.clicked.connect(self.close_report)
        header_layout.addWidget(self.close_btn)
        
        main_layout.addWidget(header_widget)

        # ==========================================
        # CONTROL BAR: FILTERS, RECONCILE & EXPORTS
        # ==========================================
        control_card = QFrame()
        control_card.setObjectName("ControlCard")
        control_card.setStyleSheet("""
            QFrame#ControlCard {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(12)
        shadow.setColor(QColor(0, 0, 0, 15))
        shadow.setOffset(0, 2)
        control_card.setGraphicsEffect(shadow)

        control_layout = QHBoxLayout(control_card)
        control_layout.setContentsMargins(16, 12, 16, 12)
        control_layout.setSpacing(12)

        # 1. Date Range Filter
        filter_lbl = QLabel("Period Filter:")
        filter_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        control_layout.addWidget(filter_lbl)

        self.date_filter_combo = QComboBox()
        self.date_filter_combo.addItems([
            "All Dates", "Q1 (Apr - Jun)", "Q2 (Jul - Sep)", "Q3 (Oct - Dec)", "Q4 (Jan - Mar)"
        ])
        self.date_filter_combo.setFixedWidth(110)
        self.date_filter_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 5px 10px;
                font-weight: 600;
                color: #1E293B;
                background-color: #F8FAFC;
            }
        """)
        self.date_filter_combo.currentTextChanged.connect(self.on_date_filter_changed)
        control_layout.addWidget(self.date_filter_combo)

        # Statement Selector
        stmt_lbl = QLabel("Statement:")
        stmt_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        control_layout.addWidget(stmt_lbl)

        self.statement_combo = QComboBox()
        self.statement_combo.setFixedWidth(160)
        self.statement_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid #CBD5E1;
                border-radius: 6px;
                padding: 5px 10px;
                font-weight: 600;
                color: #1E293B;
                background-color: #F8FAFC;
            }
        """)
        self.statement_combo.currentTextChanged.connect(self.on_statement_combo_changed)
        control_layout.addWidget(self.statement_combo)

        # 2. Reconcile GSTR-2B Button (Purple accent)
        self.gstr2b_btn = QPushButton("Import GSTR-2B / Purchase (.xlsx)")
        self.gstr2b_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.gstr2b_btn.setStyleSheet("""
            QPushButton {
                background-color: #F5F3FF;
                color: #7C3AED;
                border: 1px solid #DDD6FE;
                border-radius: 6px;
                padding: 7px 14px;
                font-weight: 600;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #EDE9FE;
            }
        """)
        self.gstr2b_btn.clicked.connect(self.import_gstr2b_file)
        control_layout.addWidget(self.gstr2b_btn)

        control_layout.addSpacing(16)
        ctrl_lbl = QLabel("Export:")
        ctrl_lbl.setStyleSheet("font-weight: 700; color: #475569; font-size: 11px; text-transform: uppercase;")
        control_layout.addWidget(ctrl_lbl)

        # Export Excel Button (Green accent)
        self.export_excel_btn = QPushButton("Excel (.xlsx)")
        self.export_excel_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #F0FDF4;
                color: #16A34A;
                border: 1px solid #BBF7D0;
                border-radius: 6px;
                padding: 7px 14px;
                font-weight: 600;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #DCFCE7; }
        """)
        self.export_excel_btn.clicked.connect(self.export_excel)
        control_layout.addWidget(self.export_excel_btn)

        # Export PDF Button (Blue accent)
        self.export_pdf_btn = QPushButton("PDF (A4 Landscape)")
        self.export_pdf_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #EFF6FF;
                color: #2563EB;
                border: 1px solid #BFDBFE;
                border-radius: 6px;
                padding: 7px 14px;
                font-weight: 600;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #DBEAFE; }
        """)
        self.export_pdf_btn.clicked.connect(self.export_pdf)
        control_layout.addWidget(self.export_pdf_btn)

        # Export CSV Button (Gray accent)
        self.export_csv_btn = QPushButton("CSV (.csv)")
        self.export_csv_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.export_csv_btn.setStyleSheet("""
            QPushButton {
                background-color: #F1F5F9;
                color: #475569;
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                padding: 7px 14px;
                font-weight: 600;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #E2E8F0; }
        """)
        self.export_csv_btn.clicked.connect(self.export_csv)
        control_layout.addWidget(self.export_csv_btn)

        # Send via Email Button (Purple/Blue accent)
        self.send_email_btn = QPushButton("✉ Send via Email")
        self.send_email_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.send_email_btn.setStyleSheet("""
            QPushButton {
                background-color: #F5F3FF;
                color: #7C3AED;
                border: 1px solid #DDD6FE;
                border-radius: 6px;
                padding: 7px 14px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #EDE9FE; }
        """)
        self.send_email_btn.clicked.connect(self.open_email_composer)
        control_layout.addWidget(self.send_email_btn)
        
        control_layout.addStretch()
        main_layout.addWidget(control_card)

        # ==========================================
        # DUAL TAB VIEW: ANALYTICS & INTERACTIVE GRID
        # ==========================================
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                background-color: #FFFFFF;
            }
            QTabBar::tab {
                background-color: #F1F5F9;
                color: #64748B;
                padding: 10px 20px;
                font-weight: 600;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                margin-right: 4px;
            }
            QTabBar::tab:selected {
                background-color: #FFFFFF;
                color: #2563EB;
                border-bottom: 2px solid #2563EB;
            }
        """)

        # Tab 1: Visual Report (QTextBrowser)
        self.report_viewer = QTextBrowser()
        self.report_viewer.setOpenExternalLinks(True)
        self.report_viewer.setStyleSheet("border: none; background-color: transparent;")
        self.tab_widget.addTab(self.report_viewer, "📊 Visual Audit & Analytics")

        # Tab 2: Interactive Grid Editor Container
        grid_container = QWidget()
        grid_container_layout = QVBoxLayout(grid_container)
        grid_container_layout.setContentsMargins(12, 12, 12, 12)
        grid_container_layout.setSpacing(8)
        
        self.grid_editor = QTableWidget()
        self.grid_editor.setAlternatingRowColors(True)
        self.grid_editor.setStyleSheet("""
            QTableWidget {
                border: none;
                gridline-color: #E2E8F0;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #FFFBEB;
                color: #B45309;
                font-weight: bold;
                border: 1px solid #FDE68A;
                padding: 6px;
            }
        """)
        self.grid_editor.cellChanged.connect(self.on_grid_cell_changed)
        grid_container_layout.addWidget(self.grid_editor)
        
        # Bottom Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.undo_btn = SecondaryButton("Undo Changes")
        self.undo_btn.setFixedWidth(130)
        self.undo_btn.clicked.connect(self.undo_changes)
        btn_layout.addWidget(self.undo_btn)
        
        self.save_btn = PrimaryButton("Save Changes")
        self.save_btn.setFixedWidth(130)
        self.save_btn.clicked.connect(self.save_changes)
        btn_layout.addWidget(self.save_btn)
        
        grid_container_layout.addLayout(btn_layout)
        self.tab_widget.addTab(grid_container, "✏ Interactive Ledger Editor")

        main_layout.addWidget(self.tab_widget, stretch=1)

    def set_active_report(self, payload, excel_path):
        """Loads and processes the payload to render the GST HTML report and populate the editor."""
        self.parsed_payload = payload
        self.excel_path = excel_path
        self.gstr2b_summary = None
        self.gstr2b_path = None
        
        transactions = payload.get("transactions", [])
        self.gst_ledger = GSTService.generate_gst_ledger(transactions)
        
        # Populate default invoice numbers and cess values
        import re
        for i, tx in enumerate(self.gst_ledger):
            if "invoice_num" not in tx or not tx["invoice_num"]:
                match = re.search(r'(?:inv|invoice|utr|no)[:\s-]*([A-Z0-9]+)', tx.get("narration", ""), re.IGNORECASE)
                if match:
                    tx["invoice_num"] = match.group(1)
                else:
                    tx["invoice_num"] = f"INV-{i+1:04d}"
            if "cess" not in tx:
                tx["cess"] = 0.0
                
        # Make a deep copy for undo
        import copy
        self.original_gst_ledger = copy.deepcopy(self.gst_ledger)
        
        # Check if GSTR-2B was previously imported for this statement
        user = UserSession.get_current_user()
        user_id = user["id"] if user else "guest"
        db = MongoDBService.get_db()
        if db is not None:
            try:
                col = db["gstr2b_imports"]
                record = col.find_one({"user_id": user_id, "excel_path": self.excel_path})
                if record and record.get("gstr2b_path") and os.path.exists(record.get("gstr2b_path")):
                    filepath = record.get("gstr2b_path")
                    self.gstr2b_path = filepath
                    # Re-run reconciliation
                    summary = GSTService.reconcile_with_gstr2b(self.gst_ledger, filepath)
                    self.gstr2b_summary = summary
            except Exception as e:
                print(f"Failed to load GSTR-2B import: {e}")
                
        self.render_all_views()
        Toast.success(self, "✓ GST Reconciliation Report & Editor rendered successfully!")

    def render_all_views(self):
        """Re-renders both the HTML report tab and the interactive QTableWidget grid."""
        if not self.parsed_payload:
            return

        # 1. Render HTML report
        html_content = GSTService.generate_gst_report_html(
            self.parsed_payload, self.gst_ledger, self.current_date_filter, self.gstr2b_summary
        )
        self.report_viewer.setHtml(html_content)

        # 2. Populate Interactive QTableWidget
        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        self.populate_grid_editor(active_ledger)

    def populate_grid_editor(self, active_ledger):
        """Populates the QTableWidget with editable fields."""
        self._is_updating_table = True
        
        headers = [
            "Invoice Number", "Invoice Date", "Supplier Name", "GSTIN", 
            "Taxable Value", "CGST", "SGST", "IGST", "CESS", "Total Tax", "Invoice Amount", "Status"
        ]
        
        self.grid_editor.clear()
        self.grid_editor.setRowCount(len(active_ledger))
        self.grid_editor.setColumnCount(len(headers))
        self.grid_editor.setHorizontalHeaderLabels(headers)
        
        for r_idx, tx in enumerate(active_ledger):
            # 0. Invoice Number (Editable)
            item_inv = QTableWidgetItem(tx.get("invoice_num", ""))
            self.grid_editor.setItem(r_idx, 0, item_inv)

            # 1. Invoice Date (Editable)
            item_date = QTableWidgetItem(tx.get("date", ""))
            self.grid_editor.setItem(r_idx, 1, item_date)

            # 2. Supplier Name (Editable)
            item_vendor = QTableWidgetItem(tx.get("vendor", ""))
            self.grid_editor.setItem(r_idx, 2, item_vendor)

            # 3. GSTIN (Editable)
            item_gstin = QTableWidgetItem(tx.get("gstin", "Unassigned"))
            self.grid_editor.setItem(r_idx, 3, item_gstin)

            # 4. Taxable Value (Editable)
            item_base = QTableWidgetItem(f"₹ {tx.get('base_value', 0.0):,.2f}")
            item_base.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 4, item_base)

            # 5. CGST (Editable)
            item_cgst = QTableWidgetItem(f"₹ {tx.get('cgst', 0.0):,.2f}")
            item_cgst.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 5, item_cgst)

            # 6. SGST (Editable)
            item_sgst = QTableWidgetItem(f"₹ {tx.get('sgst', 0.0):,.2f}")
            item_sgst.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 6, item_sgst)

            # 7. IGST (Editable)
            item_igst = QTableWidgetItem(f"₹ {tx.get('igst', 0.0):,.2f}")
            item_igst.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 7, item_igst)

            # 8. CESS (Editable)
            item_cess = QTableWidgetItem(f"₹ {tx.get('cess', 0.0):,.2f}")
            item_cess.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 8, item_cess)

            # 9. Total Tax (Editable)
            item_gst = QTableWidgetItem(f"₹ {tx.get('total_gst', 0.0):,.2f}")
            item_gst.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 9, item_gst)

            # 10. Invoice Amount (Editable)
            item_amt = QTableWidgetItem(f"₹ {tx.get('total_amount', 0.0):,.2f}")
            item_amt.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.grid_editor.setItem(r_idx, 10, item_amt)

            # 11. Status (Read-only)
            item_status = QTableWidgetItem(tx.get("gstr2b_status", "Not Reconciled"))
            item_status.setFlags(item_status.flags() & ~Qt.ItemFlag.ItemIsEditable)
            item_status.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.grid_editor.setItem(r_idx, 11, item_status)

        self.grid_editor.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.grid_editor.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._is_updating_table = False

    def on_grid_cell_changed(self, row, col):
        """Fires when any cell text is edited in the grid."""
        if self._is_updating_table or not self.gst_ledger:
            return

        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        if row < 0 or row >= len(active_ledger):
            return

        tx = active_ledger[row]
        item = self.grid_editor.item(row, col)
        if not item:
            return
            
        new_val = item.text().strip()
        clean_val = new_val.replace("₹", "").replace(",", "").strip()

        if col == 0:
            tx["invoice_num"] = new_val
        elif col == 1:
            tx["date"] = new_val
        elif col == 2:
            tx["vendor"] = new_val
        elif col == 3:
            tx["gstin"] = new_val
        elif col == 4:
            try: tx["base_value"] = float(clean_val)
            except: pass
        elif col == 5:
            try: tx["cgst"] = float(clean_val)
            except: pass
        elif col == 6:
            try: tx["sgst"] = float(clean_val)
            except: pass
        elif col == 7:
            try: tx["igst"] = float(clean_val)
            except: pass
        elif col == 8:
            try: tx["cess"] = float(clean_val)
            except: pass
        elif col == 9:
            try: tx["total_gst"] = float(clean_val)
            except: pass
        elif col == 10:
            try: tx["total_amount"] = float(clean_val)
            except: pass

        # Re-calculate Totals dynamically
        tx["total_gst"] = (tx.get("cgst") or 0.0) + (tx.get("sgst") or 0.0) + (tx.get("igst") or 0.0) + (tx.get("cess") or 0.0)
        tx["total_amount"] = (tx.get("base_value") or 0.0) + tx["total_gst"]

        # Temporarily block signals to update the calculated fields in the table
        self._is_updating_table = True
        self.grid_editor.item(row, 9).setText(f"₹ {tx['total_gst']:,.2f}")
        self.grid_editor.item(row, 10).setText(f"₹ {tx['total_amount']:,.2f}")
        self._is_updating_table = False

        # Re-render HTML view silently
        html_content = GSTService.generate_gst_report_html(
            self.parsed_payload, self.gst_ledger, self.current_date_filter, self.gstr2b_summary
        )
        self.report_viewer.setHtml(html_content)

    def on_category_edited(self, row, new_category):
        """Fires when the user changes a transaction category dropdown."""
        if self._is_updating_table or not self.gst_ledger:
            return
            
        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        if row < 0 or row >= len(active_ledger):
            return

        tx = active_ledger[row]
        tx["category"] = new_category
        tx["is_business"] = False if new_category == "Personal" else True
        
        # Recalculate rate and breakdown
        rate = GSTService.detect_gst_rate(new_category, tx["narration"])
        tx["gst_rate"] = rate
        itc_eligible = GSTService.is_itc_eligible(new_category, tx["is_business"])
        
        breakdown = GSTService.calculate_gst_breakdown(tx["total_amount"], rate, tx["narration"])
        tx["base_value"] = breakdown["base_value"]
        tx["cgst"] = breakdown["cgst"]
        tx["sgst"] = breakdown["sgst"]
        tx["igst"] = breakdown["igst"]
        tx["total_gst"] = breakdown["total_gst"]
        tx["itc_eligible"] = "Yes" if (itc_eligible and breakdown["total_gst"] > 0) else "No"
        
        self.render_all_views()
        Toast.success(self, f"Updated row #{row+1} category to '{new_category}'")

    def on_itc_edited(self, row, new_itc):
        """Fires when the user toggles ITC eligibility."""
        if self._is_updating_table or not self.gst_ledger:
            return
            
        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        if row < 0 or row >= len(active_ledger):
            return

        tx = active_ledger[row]
        tx["itc_eligible"] = new_itc
        
        html_content = GSTService.generate_gst_report_html(
            self.parsed_payload, self.gst_ledger, self.current_date_filter, self.gstr2b_summary
        )
        self.report_viewer.setHtml(html_content)

    def on_date_filter_changed(self, new_filter):
        """Fires when user selects a new period filter dropdown option."""
        self.current_date_filter = new_filter
        self.render_all_views()
        Toast.info(self, f"Filtered view to {new_filter}")

    def import_gstr2b_file(self):
        """Prompts user to select a GSTR-2B or Purchase Register file for 3-way reconciliation."""
        default_dir = os.path.dirname(self.excel_path) if self.excel_path else ""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Select GSTR-2B or Purchase Register File", default_dir, "Excel / CSV Files (*.xlsx *.xls *.csv)"
        )
        if not filepath:
            return

        # Show status loading/info toast
        Toast.info(self, "Running background GST Reconciliation...")
        
        self.reconcile_worker = GSTR2BReconcileWorker(self.gst_ledger, filepath, self)
        
        def on_finished(summary):
            self.gstr2b_summary = summary
            self.gstr2b_path = filepath
            
            # Store GSTR-2B in user's account
            user = UserSession.get_current_user()
            user_id = user["id"] if user else "guest"
            db = MongoDBService.get_db()
            if db is not None:
                try:
                    col = db["gstr2b_imports"]
                    col.update_one(
                        {"user_id": user_id, "excel_path": self.excel_path},
                        {"$set": {
                            "gstr2b_path": filepath,
                            "import_date": datetime.datetime.utcnow().isoformat()
                        }},
                        upsert=True
                    )
                except Exception as e:
                    print(f"Failed to store GSTR-2B path in db: {e}")
                    
            self.render_all_views()
            
            QMessageBox.information(
                self, "3-Way Reconciliation Complete",
                f"Successfully matched bank transactions against GSTR-2B file!\n\n"
                f"• Matched Entries: {summary['matched_count']}\n"
                f"• Missing in GSTR-2B: {summary['missing_count']}\n"
                f"• Amount Discrepancies: {summary['discrepancy_count']}\n"
                f"• Reconciled ITC: ₹ {summary['matched_gst']:,.2f}"
            )
            self.reconcile_worker.deleteLater()
            self.reconcile_worker = None
            
        def on_error(err):
            QMessageBox.critical(self, "Reconciliation Error", f"Could not process GSTR-2B file:\n{err}")
            self.reconcile_worker.deleteLater()
            self.reconcile_worker = None
            
        self.reconcile_worker.finished.connect(on_finished)
        self.reconcile_worker.error.connect(on_error)
        self.reconcile_worker.start()

    def export_excel(self):
        """Generates and saves the updated GST Excel report."""
        if not self.gst_ledger:
            return
            
        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        default_dir = os.path.dirname(self.excel_path) if self.excel_path else os.path.expanduser("~/Documents")
        bank = self.parsed_payload.get("bank_name", "Bank") if self.parsed_payload else "Statement"
        filename = f"GST_Reconciliation_{bank}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        
        dest_path, _ = QFileDialog.getSaveFileName(
            self, "Save GST Excel Report", os.path.join(default_dir, filename), "Excel Files (*.xlsx)"
        )
        if not dest_path:
            return

        try:
            from parser.gst_excel_writer import GSTExcelWriter
            out_path = GSTExcelWriter.write_gst_excel(
                dest_path,
                self.parsed_payload.get("bank_name", "Bank"),
                self.parsed_payload.get("account_holder", "Holder"),
                self.parsed_payload.get("period", "Period"),
                active_ledger
            )
            Toast.success(self, "✓ Excel Ledger exported successfully!")
            
            if os.name == 'nt':
                os.startfile(out_path)
            else:
                import subprocess
                subprocess.run(["open", out_path] if os.name == 'posix' else ["xdg-open", out_path])
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not write Excel file:\n{e}")

    def export_pdf(self):
        """Prints the report HTML to a high-quality A4 Landscape PDF."""
        html_content = self.report_viewer.toHtml()
        if not html_content:
            return

        bank = self.parsed_payload.get("bank_name", "Bank") if self.parsed_payload else "Statement"
        filename = f"GST_Reconciliation_Report_{bank}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
        default_dir = os.path.dirname(self.excel_path) if self.excel_path else os.path.expanduser("~/Documents")
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Save GST PDF Report", os.path.join(default_dir, filename), "PDF Files (*.pdf)"
        )
        if not filepath:
            return

        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(filepath)
            
            printer.setPageLayout(
                QPageLayout(
                    QPageSize(QPageSize.PageSizeId.A4),
                    QPageLayout.Orientation.Landscape,
                    QMarginsF(10, 10, 10, 10),
                    QPageLayout.Unit.Millimeter
                )
            )
            
            doc = QTextDocument()
            doc.setHtml(html_content)
            doc.print_(printer)
            
            Toast.success(self, "✓ PDF Report printed successfully in Landscape!")
            
            if os.path.exists(filepath):
                if os.name == 'nt':
                    os.startfile(filepath)
                else:
                    import subprocess
                    subprocess.run(["open", filepath] if os.name == 'posix' else ["xdg-open", filepath])
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not print A4 PDF:\n{e}")

    def export_csv(self):
        """Generates and saves the GST transactions ledger to a CSV file."""
        if not self.gst_ledger:
            return

        active_ledger = GSTService.filter_ledger_by_date(self.gst_ledger, self.current_date_filter)
        bank = self.parsed_payload.get("bank_name", "Bank") if self.parsed_payload else "Statement"
        filename = f"GST_Ledger_{bank}_{datetime.datetime.now().strftime('%Y%m%d')}.csv"
        default_dir = os.path.dirname(self.excel_path) if self.excel_path else os.path.expanduser("~/Documents")
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Save GST CSV Ledger", os.path.join(default_dir, filename), "CSV Files (*.csv)"
        )
        if not filepath:
            return

        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "Date", "Narration", "Category", "Vendor Name", "Vendor GSTIN", 
                    "Total Amount", "Taxable Value", "GST Rate", "CGST", "SGST", "IGST", 
                    "Total GST", "ITC Eligible", "GSTR-2B Status", "AI Confidence", "Status"
                ])
                
                for tx in active_ledger:
                    writer.writerow([
                        tx.get("date", ""),
                        tx.get("narration", ""),
                        tx.get("category", ""),
                        tx.get("vendor", ""),
                        tx.get("gstin", "Unassigned"),
                        tx.get("total_amount", 0.0),
                        tx.get("base_value", 0.0),
                        f"{tx.get('gst_rate', 0.18)*100:.0f}%",
                        tx.get("cgst", 0.0),
                        tx.get("sgst", 0.0),
                        tx.get("igst", 0.0),
                        tx.get("total_gst", 0.0),
                        tx.get("itc_eligible", "No"),
                        tx.get("gstr2b_status", "Not Reconciled"),
                        f"{tx.get('confidence', 80):.0f}%",
                        tx.get("status", "Estimated")
                    ])
            
            Toast.success(self, "✓ GST CSV Ledger exported successfully!")
            
            if os.name == 'nt':
                os.startfile(filepath)
            else:
                import subprocess
                subprocess.run(["open", filepath] if os.name == 'posix' else ["xdg-open", filepath])
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not save CSV ledger:\n{e}")

    def open_email_composer(self):
        """Opens the Email Composer dialog pre-attaching the GST report."""
        from ui.email_composer_dialog import EmailComposerDialog
        
        attachment = getattr(self, "excel_path", None)
        period = self.current_date_filter
        bank = self.parsed_payload.get("bank_name", "") if self.parsed_payload else "Statement"

        dialog = EmailComposerDialog(
            report_type="GST Reconciliation & Analysis Report",
            default_attachment=attachment,
            period=period,
            bank_name=bank,
            parent=self
        )
        dialog.exec()

    def close_report(self):
        """Clears states and returns to the dashboard screen."""
        self.parsed_payload = None
        self.gst_ledger = []
        self.excel_path = None
        self.gstr2b_summary = None
        self.report_viewer.clear()
        
        self.closed.emit()
        
        p = self.parent()
        while p:
            if hasattr(p, "switch_dashboard_page"):
                p.switch_dashboard_page("dashboard")
                break
            p = p.parent()

    def update_theme_style(self, theme):
        """Updates colors and themes dynamically to match app settings."""
        self.current_theme = theme
        if theme == "dark":
            self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #F8FAFC;")
            self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #94A3B8;")
            self.findChild(QFrame, "ControlCard").setStyleSheet("""
                QFrame#ControlCard {
                    background-color: #1E293B;
                    border: 1px solid #334155;
                    border-radius: 12px;
                }
            """)
        else:
            self.title_lbl.setStyleSheet("font-size: 22px; font-weight: 700; color: #0F172A;")
            self.subtitle_lbl.setStyleSheet("font-size: 13px; color: #64748B;")
            self.findChild(QFrame, "ControlCard").setStyleSheet("""
                QFrame#ControlCard {
                    background-color: #FFFFFF;
                    border: 1px solid #E2E8F0;
                    border-radius: 12px;
                }
            """)

    def _safe_run_query(self, query_fn, callback_fn):
        worker = DBQueryWorker(query_fn, self)
        if not hasattr(self, "_active_workers"):
            self._active_workers = []
        self._active_workers.append(worker)
        
        def handle_result(res):
            try:
                callback_fn(res)
            finally:
                if worker in self._active_workers:
                    self._active_workers.remove(worker)
                worker.deleteLater()
                
        worker.result_ready.connect(handle_result)
        worker.start()

    def load_statements_dropdown(self):
        """Loads all completed GST statements for the current user into the dropdown."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else "guest"
        
        def db_query():
            col = MongoDBService.get_collection()
            if col is not None:
                return list(col.find(
                    {"user_id": user_id, "status": "Completed", "output_format": "GST Report"},
                    sort=[("upload_date", -1)]
                ))
            return []
            
        def db_callback(logs):
            self.statement_combo.blockSignals(True)
            self.statement_combo.clear()
            self.statement_combo.addItem("Select a GST Statement...", "")
            
            self._statement_paths = {}
            for log in logs:
                excel_path = log.get("excel_path")
                if excel_path and os.path.exists(excel_path):
                    filename = os.path.basename(log.get("pdf_path", "Statement.pdf"))
                    upload_date = log.get("upload_date")
                    if hasattr(upload_date, "strftime"):
                        date_str = upload_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(upload_date or "")[:10]
                    display_text = f"{log.get('bank_name', 'Bank')} ({date_str}) - {filename}"
                    self.statement_combo.addItem(display_text, excel_path)
                    self._statement_paths[display_text] = excel_path
            self.statement_combo.blockSignals(False)
            
            # Automatically load the first statement if available
            if self.statement_combo.count() > 1:
                self.statement_combo.setCurrentIndex(1)
                
        self._safe_run_query(db_query, db_callback)

    def on_statement_combo_changed(self, display_text):
        """Loads the selected statement from the dropdown."""
        if not display_text or not hasattr(self, "_statement_paths"):
            return
            
        excel_path = self._statement_paths.get(display_text)
        if excel_path:
            self.on_statement_selected_by_path(excel_path)

    def on_statement_selected_by_path(self, excel_path):
        """Loads and processes details from the selected Excel path."""
        if not excel_path:
            return
            
        # Fallback to standard Excel if this is a GST Report and the standard one exists
        if "_GST_Report.xlsx" in excel_path:
            possible_std = excel_path.replace("_GST_Report.xlsx", ".xlsx")
            if os.path.exists(possible_std):
                excel_path = possible_std
                
        if not os.path.exists(excel_path):
            return
            
        try:
            # Read transactions from excel workbook
            transactions = self.load_transactions_from_excel(excel_path)
            meta = self.load_summary_from_excel(excel_path)
            
            payload = {
                "transactions": transactions,
                "bank_name": meta.get("bank_name", "Unknown Bank"),
                "account_holder": meta.get("account_holder", "Unknown"),
                "account_number": meta.get("account_number", "Unknown"),
                "period": meta.get("period", "Unknown Period"),
                "currency": meta.get("currency", "INR")
            }
            
            self.set_active_report(payload, excel_path)
        except Exception as e:
            QMessageBox.critical(self, "Error Loading Excel", f"Could not load transaction sheets from Excel archive:\n{e}")

    def load_transactions_from_excel(self, excel_path) -> list:
        """Helper to parse raw transaction ledger details from processed Excel outputs."""
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        transactions = []
        
        # Try loading standard transactions sheet
        sheet_names = wb.sheetnames
        target_sheet = None
        for name in ("Transactions Ledger", "GST Transactions", "Sheet1"):
            if name in sheet_names:
                target_sheet = wb[name]
                break
        if not target_sheet:
            target_sheet = wb.active
            
        rows = list(target_sheet.iter_rows(values_only=True))
        if not rows:
            return []
            
        # Inspect headers
        header_row = rows[0]
        col_map = {}
        for idx, h in enumerate(header_row):
            if h:
                col_map[str(h).strip().lower()] = idx
                
        for row in rows[1:]:
            if not any(row):
                continue
                
            date_idx = col_map.get("date") or col_map.get("transaction date")
            narr_idx = col_map.get("narration") or col_map.get("description") or col_map.get("particulars")
            val_idx = col_map.get("value") or col_map.get("amount") or col_map.get("total amount")
            type_idx = col_map.get("type")
            debit_idx = col_map.get("debit")
            credit_idx = col_map.get("credit")
            bal_idx = col_map.get("balance")
            
            tx = {}
            if date_idx is not None and date_idx < len(row):
                tx["date"] = str(row[date_idx]) if row[date_idx] is not None else ""
            if narr_idx is not None and narr_idx < len(row):
                tx["narration"] = str(row[narr_idx]) if row[narr_idx] is not None else ""
                
            # Debit/Credit
            debit_val = 0.0
            credit_val = 0.0
            if debit_idx is not None and debit_idx < len(row):
                try: debit_val = float(str(row[debit_idx]).replace(",", "").strip())
                except: pass
            if credit_idx is not None and credit_idx < len(row):
                try: credit_val = float(str(row[credit_idx]).replace(",", "").strip())
                except: pass
                
            if val_idx is not None and val_idx < len(row) and row[val_idx] is not None:
                try:
                    val = float(str(row[val_idx]).replace(",", "").strip())
                    if type_idx is not None and type_idx < len(row):
                        t_str = str(row[type_idx]).lower()
                        if "debit" in t_str:
                            debit_val = val
                        else:
                            credit_val = val
                    else:
                        if val < 0:
                            debit_val = abs(val)
                        else:
                            credit_val = val
                except:
                    pass
                    
            tx["debit"] = debit_val
            tx["credit"] = credit_val
            
            # balance
            if bal_idx is not None and bal_idx < len(row):
                try: tx["balance"] = float(str(row[bal_idx]).replace(",", "").strip())
                except: tx["balance"] = 0.0
            else:
                tx["balance"] = 0.0
                
            # Default type derivation if missing
            if type_idx is not None and type_idx < len(row) and row[type_idx]:
                tx["type"] = str(row[type_idx]).strip()
            else:
                tx["type"] = "Debit" if debit_val > 0 else "Credit"
                
            # Copy other keys if present
            for k in ("category", "vendor", "gstin", "base_value", "total_gst", "cgst", "sgst", "igst", "cess", "invoice_num", "itc_eligible", "gstr2b_status"):
                k_idx = col_map.get(k)
                if k_idx is not None and k_idx < len(row) and row[k_idx] is not None:
                    if k in ("base_value", "total_gst", "cgst", "sgst", "igst", "cess"):
                        try: tx[k] = float(str(row[k_idx]).replace(",", "").strip())
                        except: tx[k] = 0.0
                    else:
                        tx[k] = str(row[k_idx]).strip()
                        
            transactions.append(tx)
        return transactions

    def load_summary_from_excel(self, excel_path) -> dict:
        """Helper to parse overview summary metadata cards from spreadsheet header regions."""
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        meta = {}
        
        # Try loading summary sheet
        sheet_names = wb.sheetnames
        target_sheet = None
        for name in ("Executive Summary", "Summary", "Sheet1"):
            if name in sheet_names:
                target_sheet = wb[name]
                break
        if not target_sheet:
            target_sheet = wb.active
            
        for r in range(1, 15):
            for c in range(1, 10):
                val = target_sheet.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    val_clean = val.strip().lower()
                    next_val = target_sheet.cell(row=r, column=c+1).value
                    if "bank name" in val_clean:
                        meta["bank_name"] = str(next_val).strip() if next_val else ""
                    elif "holder" in val_clean or "client name" in val_clean:
                        meta["account_holder"] = str(next_val).strip() if next_val else ""
                    elif "account number" in val_clean:
                        meta["account_number"] = str(next_val).strip() if next_val else ""
                    elif "period" in val_clean:
                        meta["period"] = str(next_val).strip() if next_val else ""
        return meta

    def undo_changes(self):
        """Restores the ledger to its original unedited state."""
        if not hasattr(self, "original_gst_ledger") or not self.original_gst_ledger:
            Toast.info(self, "No changes to undo")
            return
            
        import copy
        self.gst_ledger = copy.deepcopy(self.original_gst_ledger)
        self.render_all_views()
        Toast.success(self, "✓ Restored original ledger states!")

    def save_changes(self):
        """Saves the edited ledger rows back to the Excel file database."""
        if not self.excel_path or not self.gst_ledger:
            Toast.warning(self, "No active statement to save")
            return
            
        try:
            from parser.gst_excel_writer import GSTExcelWriter
            # Save to Excel
            GSTExcelWriter.write_gst_excel(
                self.excel_path,
                self.parsed_payload.get("bank_name", "Bank"),
                self.parsed_payload.get("account_holder", "Holder"),
                self.parsed_payload.get("period", "Period"),
                self.gst_ledger
            )
            
            import copy
            self.original_gst_ledger = copy.deepcopy(self.gst_ledger)
            
            # Recalculate reconciliation metrics if GSTR-2B was loaded
            if self.gstr2b_summary and getattr(self, "gstr2b_path", None):
                summary = GSTService.reconcile_with_gstr2b(self.gst_ledger, self.gstr2b_path)
                self.gstr2b_summary = summary
                
            self.render_all_views()
            Toast.success(self, "✓ Changes saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Save Failed", f"Could not save changes to Excel database:\n{e}")
