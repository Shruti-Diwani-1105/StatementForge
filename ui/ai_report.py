import os
import datetime
import json
import re
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QMessageBox, QFileDialog
)
from PyQt6.QtCore import pyqtSignal, QThread, QSize, QTimer
from PyQt6.QtGui import QTextDocument
from PyQt6.QtPrintSupport import QPrinter

from ui.html_screen_wrapper import HtmlScreenWrapper
from services.gemini_service import GeminiService
from services.history_service import HistoryService
from utils.user_session import UserSession
from settings.toast import Toast

def format_indian_currency(amount: float) -> str:
    """Formats a float value into Indian currency format (e.g., ₹ 15,42,38,560.25)."""
    try:
        is_negative = amount < 0
        amount = abs(amount)
        
        s = f"{amount:.2f}"
        parts = s.split(".")
        int_part = parts[0]
        dec_part = parts[1]
        
        if len(int_part) <= 3:
            res = int_part
        else:
            thousands_part = int_part[-3:]
            remaining = int_part[:-3]
            pairs = []
            while len(remaining) > 0:
                pairs.append(remaining[-2:])
                remaining = remaining[:-2]
            pairs.reverse()
            res = ",".join(pairs) + "," + thousands_part
            
        formatted = f"₹ {res}.{dec_part}"
        return f"-{formatted}" if is_negative else formatted
    except Exception:
        return f"₹ {amount:,.2f}"


class DBQueryWorker(QThread):
    result_ready = pyqtSignal(object)

    def __init__(self, query_fn, parent=None):
        super().__init__(parent)
        self.query_fn = query_fn

    def run(self):
        try:
            res = self.query_fn()
            self.result_ready.emit(res)
        except Exception as e:
            print(f"DBQueryWorker error: {e}")
            self.result_ready.emit(None)


class AIWorker(QThread):
    """
    Background worker thread to execute Gemini analysis methods without blocking the UI.
    """
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, action, transactions, currency="INR", **kwargs):
        super().__init__()
        self.action = action
        self.transactions = transactions
        self.currency = currency
        self.kwargs = kwargs

    def run(self):
        try:
            if not self.transactions:
                raise ValueError("No transaction data loaded. Please upload or select a bank statement first.")

            bank_name = self.kwargs.get("bank_name", "Unknown Bank")
            period = self.kwargs.get("period", "Unknown Period")

            if self.action == "summary":
                result = GeminiService.generate_financial_summary(
                    self.transactions, bank_name, period, self.currency
                )
            elif self.action == "spending":
                result = GeminiService.analyze_monthly_spending(
                    self.transactions, self.currency, bank_name=bank_name, period=period
                )
            elif self.action == "risk":
                result = GeminiService.analyze_risks(
                    self.transactions, self.currency, bank_name=bank_name, period=period
                )
            elif self.action == "report":
                holder = self.kwargs.get("holder", "Unknown")
                acc_num = self.kwargs.get("acc_num", "Unknown")
                result = GeminiService.generate_executive_report(
                    self.transactions, bank_name, holder, acc_num, period, self.currency
                )
            else:
                raise ValueError(f"Unknown AI action: {self.action}")

            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class PrepareAllReportsWorker(QThread):
    """
    Background worker thread that renders all 4 financial reports
    once per statement selection from the single source of truth report_data.
    """
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, transactions, currency="INR", **kwargs):
        super().__init__()
        self.transactions = transactions
        self.currency = currency
        self.kwargs = kwargs

    def run(self):
        try:
            if not self.transactions:
                raise ValueError("No sufficient transaction data available.")

            bank_name = self.kwargs.get("bank_name", "Unknown Bank")
            period = self.kwargs.get("period", "Unknown Period")
            holder = self.kwargs.get("holder", "Unknown")
            acc_num = self.kwargs.get("acc_num", "Unknown")

            report_data = GeminiService.build_report_data(
                self.transactions,
                bank_name,
                period,
                account_holder=holder,
                account_number=acc_num,
                currency=self.currency
            )

            # Render 4 distinct UI views
            summary_html = GeminiService._render_financial_summary_view(report_data)
            spending_html = GeminiService._render_spending_insights_view(report_data)
            risk_html = GeminiService._render_risk_analysis_view(report_data)
            full_report_html = GeminiService._render_full_report_view(report_data)

            results = {
                "summary": summary_html,
                "spending": spending_html,
                "risk": risk_html,
                "report": full_report_html
            }

            self.finished.emit(results)
        except Exception as e:
            self.error.emit(str(e))


class AIReportWidget(QWidget):
    """
    Main UI section for AI Financial Report feature powered by HTML + CSS presentation layer.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_theme = "light"
        
        # State Data & Single Source of Truth Report Cache
        self.active_transactions = []
        self.report_data = {}
        self.active_metadata = {
            "bank_name": "Unknown Bank",
            "account_holder": "Unknown",
            "period": "Unknown Period",
            "currency": "INR",
            "total_credit": 0.0,
            "total_debit": 0.0,
            "net_savings": 0.0
        }
        self.prepared_reports = {}
        self.reports_ready = False
        self.active_action_key = "summary"
        self.prepare_thread = None
        self.active_thread = None
        self.current_report_html = ""

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.html_wrapper = HtmlScreenWrapper("web/ai_report.html", self)
        layout.addWidget(self.html_wrapper)

        self.html_wrapper.web_view.titleChanged.connect(self.handle_web_commands)
        QTimer.singleShot(600, self.load_history_dropdown)

    def update_theme_style(self, theme: str = "light"):
        """Updates HTML UI theme styling ('light' or 'dark')."""
        self.current_theme = theme.lower().strip() if isinstance(theme, str) else "light"
        if self.current_theme == "dark":
            self.html_wrapper.eval_js("document.body.classList.add('dark-mode');")
        else:
            self.html_wrapper.eval_js("document.body.classList.remove('dark-mode');")

    def handle_web_commands(self, title: str):
        """Dispatches commands sent from JavaScript UI via document.title IPC."""
        if not title or not title.startswith("app-cmd:"):
            return
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, lambda: self._process_web_commands(title))

    def _process_web_commands(self, title: str):
        parts = title.split(":", 2)
        cmd = parts[1] if len(parts) > 1 else ""
        raw_payload = parts[2] if len(parts) > 2 else ""

        if cmd == "auditor_refresh_history":
            self.load_history_dropdown()
        elif cmd == "auditor_select_statement":
            try:
                data = json.loads(raw_payload)
                excel_path = data.get("path")
                if excel_path:
                    self.on_statement_selected_by_path(excel_path)
            except Exception as e:
                print(f"select error: {e}")
        elif cmd == "auditor_run_task":
            try:
                data = json.loads(raw_payload)
                action_key = data.get("action", "summary")
                self.run_ai_task(action_key)
            except Exception:
                pass
        elif cmd == "auditor_send_email":
            self.open_email_composer()
        elif cmd == "auditor_export_pdf":
            self.export_pdf_report()

    def set_active_statement(self, payload):
        """Allows direct programmatic insertion of parsed statements."""
        if not payload or not payload.get("transactions"):
            return

        raw_txs = payload.get("transactions", [])
        self.active_transactions = GeminiService.normalize_transactions(raw_txs)
        
        bank_name = payload.get("bank_name", "Unknown Bank")
        account_holder = payload.get("account_holder", "Unknown")
        account_number = payload.get("account_number", "Unknown")
        period = payload.get("period", "Unknown Period")
        currency = payload.get("currency", "INR")

        self.report_data = GeminiService.build_report_data(
            self.active_transactions,
            bank_name,
            period,
            account_holder=account_holder,
            account_number=account_number,
            currency=currency
        )
        
        self.active_metadata = {
            "bank_name": bank_name,
            "account_holder": account_holder,
            "account_number": account_number,
            "period": period,
            "currency": currency,
            "total_credit": self.report_data["total_credits"],
            "total_debit": self.report_data["total_debits"],
            "net_savings": self.report_data["net_savings"]
        }

        # Update metrics preview in HTML top bar
        self.update_metrics_ui()
        
        # Reset analysis cache for new statement selection
        self.prepared_reports = {}
        self.reports_ready = False
        self.active_action_key = "summary"

        # Prepare all 4 report analysis outputs in background once
        self.start_prepare_all_reports()

    def start_prepare_all_reports(self):
        """Prepares and caches all 4 reports once per statement selection in background."""
        if not self.active_transactions:
            return

        if hasattr(self, "prepare_thread") and self.prepare_thread is not None:
            if self.prepare_thread.isRunning():
                try:
                    self.prepare_thread.finished.disconnect()
                    self.prepare_thread.error.disconnect()
                except Exception:
                    pass
                self.prepare_thread.terminate()
                self.prepare_thread.wait(300)
            self.prepare_thread = None

        loading_html = "<div style=\"color:#3B82F6; font-family: Inter, sans-serif; font-size:14px; text-align:center; padding-top:40px;\"><b>Preparing financial insights...</b><br>Analyzing transaction data for all 4 report modules...</div>"
        self.html_wrapper.eval_js("setLoading(true, 'Preparing financial insights...');")
        self.html_wrapper.eval_js(f"setReportHtml({json.dumps(loading_html)});")

        kwargs = {
            "bank_name": self.active_metadata.get("bank_name", "Unknown Bank"),
            "holder": self.active_metadata.get("account_holder", "Unknown"),
            "acc_num": self.active_metadata.get("account_number", "Unknown"),
            "period": self.active_metadata.get("period", "Unknown Period")
        }

        self.prepare_thread = PrepareAllReportsWorker(
            self.active_transactions,
            self.active_metadata.get("currency", "INR"),
            **kwargs
        )

        def handle_finished(results):
            self.html_wrapper.eval_js("setLoading(false);")
            self.prepared_reports = results
            self.reports_ready = True
            self.prepare_thread = None

            # Render currently selected tab report instantly
            active_html = self.prepared_reports.get(self.active_action_key, "")
            if active_html:
                active_html = GeminiService.clean_html_response(active_html)
                self.current_report_html = active_html
                self.html_wrapper.eval_js(f"setReportHtml({json.dumps(active_html)});")
                match = re.search(r"(\d{2,3})\s*/\s*100", active_html)
                if match:
                    score_val = match.group(1)
                    self.html_wrapper.eval_js(f"document.getElementById('lblScore').innerText = {json.dumps(score_val)};")

            Toast.success(self, "✓ Analysis Ready")

        def handle_error(err_msg):
            self.html_wrapper.eval_js("setLoading(false);")
            self.prepare_thread = None
            error_html = f"<div style='color:#EF4444; font-family: \"Inter\", sans-serif; font-size:14px; padding:20px;'><b>AI Report Preparation Failed</b><br><br>{err_msg}</div>"
            self.html_wrapper.eval_js(f"setReportHtml({json.dumps(error_html)});")

        self.prepare_thread.finished.connect(handle_finished)
        self.prepare_thread.error.connect(handle_error)
        self.prepare_thread.start()

    def update_metrics_ui(self):
        """Fills UI metric values from single source of truth report_data into HTML."""
        if hasattr(self, "report_data") and self.report_data:
            bank = self.report_data.get("bank_name", "-")
            period = self.report_data.get("statement_period", "-")
            total_credits = self.report_data.get("total_credits", 0.0)
            total_debits = self.report_data.get("total_debits", 0.0)
            net = self.report_data.get("net_savings", 0.0)
            score_num = self.report_data.get("risk_analysis", {}).get("score", 95)
        else:
            bank = self.active_metadata.get("bank_name", "-")
            period = self.active_metadata.get("period", "-")
            total_credits = self.active_metadata.get("total_credit", 0.0)
            total_debits = self.active_metadata.get("total_debit", 0.0)
            net = self.active_metadata.get("net_savings", 0.0)
            score_num = 95

        credits_str = format_indian_currency(total_credits)
        debits_str = format_indian_currency(total_debits)
        savings_str = format_indian_currency(net)
        is_positive = net >= 0
        score_str = f"{score_num}%" if len(self.active_transactions) > 0 else "-"

        js_code = (
            f"setMetrics("
            f"{json.dumps(bank)}, "
            f"{json.dumps(period)}, "
            f"{json.dumps(credits_str)}, "
            f"{json.dumps(debits_str)}, "
            f"{json.dumps(savings_str)}, "
            f"{json.dumps(score_str)}, "
            f"{'true' if is_positive else 'false'});"
        )
        self.html_wrapper.eval_js(js_code)

    def _safe_run_query(self, query_fn, callback_fn):
        worker = DBQueryWorker(query_fn, self)
        if not hasattr(self, "_active_workers"):
            self._active_workers = []
        self._active_workers.append(worker)
        
        def handle_result(res):
            if worker in self._active_workers:
                self._active_workers.remove(worker)
            if res is not None:
                callback_fn(res)
                
        worker.result_ready.connect(handle_result)
        worker.start()

    def load_history_dropdown(self):
        """Queries local history database logs and updates dropdown list in HTML."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else "guest"
        
        def db_query():
            return HistoryService.get_history_logs(user_id=user_id)
            
        def db_callback(logs):
            completed_logs = [log for log in logs if log.get("status") == "Completed" and log.get("excel_path")]
            
            options_js = "(function(){ var cb = document.getElementById('statementCb'); if(!cb) return; cb.innerHTML = '';"
            if not completed_logs:
                options_js += "cb.innerHTML += `<option value=''>No parsed statements found in history.</option>`;"
            else:
                options_js += "cb.innerHTML += `<option value=''>Select from parsed statement history...</option>`;"
                for log in completed_logs:
                    pdf_path = log.get("pdf_path", "")
                    excel_path = log.get("excel_path", "")
                    filename = os.path.basename(pdf_path) if pdf_path else "Statement.pdf"
                    upload_date = log.get("upload_date")
                    if hasattr(upload_date, "strftime"):
                        date_str = upload_date.strftime("%Y-%m-%d")
                    elif isinstance(upload_date, str):
                        date_str = upload_date[:10]
                    else:
                        date_str = str(upload_date or "")[:10]
                    bank = log.get("bank_name", "Unknown Bank")
                    display_text = f"{bank} ({date_str}) - {filename}"
                    escaped_path = json.dumps(excel_path)
                    options_js += f"cb.innerHTML += `<option value={escaped_path}>{display_text}</option>`;"
            options_js += "})();"
            self.html_wrapper.eval_js(options_js)

        self._safe_run_query(db_query, db_callback)

    def on_statement_selected_by_path(self, excel_path):
        """Loads and processes transaction details from history Excel path when selected in HTML."""
        if not excel_path:
            return
            
        if "_GST_Report.xlsx" in excel_path:
            possible_std = excel_path.replace("_GST_Report.xlsx", ".xlsx")
            if os.path.exists(possible_std):
                excel_path = possible_std
                
        if not os.path.exists(excel_path):
            return
            
        try:
            self.active_excel_path = excel_path
            transactions = self.load_transactions_from_excel(excel_path)
            meta = self.load_summary_from_excel(excel_path)
            
            # Statement Diagnostic Output
            total_c = sum(t.get("credit", 0.0) for t in transactions)
            total_d = sum(t.get("debit", 0.0) for t in transactions)
            print("=== SELECTED STATEMENT DIAGNOSTIC ===")
            print(f"Statement ID / Path: {excel_path}")
            print(f"Bank Name: {meta.get('bank_name', 'Unknown Bank')}")
            print(f"Period: {meta.get('period', 'Unknown Period')}")
            print(f"Transaction Count: {len(transactions)}")
            print(f"Total Credits: INR {total_c:,.2f}")
            print(f"Total Debits: INR {total_d:,.2f}")

            if not transactions:
                err_html = "<div class='report-container' style='text-align:center; padding:40px; color:#EF4444;'><b>Unable to load transaction data for the selected statement.</b><br>Please refresh the statement list or upload a valid bank statement file.</div>"
                self.html_wrapper.eval_js(f"setReportHtml({json.dumps(err_html)});")
                Toast.error(self, "Unable to load transaction data for selected statement.")
                return

            payload = {
                "transactions": transactions,
                "bank_name": meta.get("bank_name", "Unknown Bank"),
                "account_holder": meta.get("account_holder", "Unknown"),
                "account_number": meta.get("account_number", "Unknown"),
                "period": meta.get("period", "Unknown Period"),
                "currency": meta.get("currency", "INR")
            }
            
            self.set_active_statement(payload)
            Toast.success(self, "✓ Statement loaded successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error Loading Excel", f"Could not load transaction sheets from Excel archive:\n{e}")
            self.load_history_dropdown()

    def load_transactions_from_excel(self, excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_name = None
        for name in ["Transactions", "GST Transactions", "Ledger", "Statement Ledger"]:
            if name in wb.sheetnames:
                sheet_name = name
                break

        if not sheet_name:
            for name in wb.sheetnames:
                if name not in ["Summary", "GST Summary", "Settings"]:
                    sheet_name = name
                    break

        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
            
        # Robust Header Row Detection (scans rows 0..4)
        header_row_idx = 0
        for idx, r in enumerate(rows[:5]):
            if not r:
                continue
            r_str = " ".join([str(c or "").lower() for c in r])
            if any(k in r_str for k in ["date", "debit", "credit", "narration", "description", "particulars"]):
                header_row_idx = idx
                break

        headers = rows[header_row_idx]
        data_rows = rows[header_row_idx + 1:]
        
        col_mapping = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            h_lower = str(header).lower().strip()
            if "date" in h_lower and "value" not in h_lower and "date" not in col_mapping:
                col_mapping["date"] = idx
            elif any(k in h_lower for k in ["description", "narration", "particulars", "details", "remark"]) and "narration" not in col_mapping:
                col_mapping["narration"] = idx
            elif any(k in h_lower for k in ["debit", "withdrawal", "dr", "outflow"]) and "debit" not in col_mapping:
                col_mapping["debit"] = idx
            elif any(k in h_lower for k in ["credit", "deposit", "cr", "inflow"]) and "credit" not in col_mapping:
                col_mapping["credit"] = idx
            elif any(k in h_lower for k in ["balance", "running"]) and "balance" not in col_mapping:
                col_mapping["balance"] = idx
            elif "type" in h_lower and "type" not in col_mapping:
                col_mapping["type"] = idx
            elif ("amount" in h_lower or "total amount" in h_lower) and "total_amount" not in col_mapping:
                col_mapping["total_amount"] = idx
                
        transactions = []
        for r in data_rows:
            tx = {}
            has_val = False
            for k in ["date", "narration", "debit", "credit", "balance", "type", "total_amount"]:
                idx = col_mapping.get(k)
                if idx is not None and idx < len(r):
                    val = r[idx]
                    if val is not None:
                        has_val = True
                        if k == "date" and hasattr(val, "strftime"):
                            tx[k] = val.strftime("%Y-%m-%d")
                        elif k in ["debit", "credit", "balance", "total_amount"]:
                            try:
                                clean_val = str(val).replace(",", "").replace("₹", "").replace("$", "").strip()
                                if clean_val.endswith("-"):
                                    clean_val = "-" + clean_val[:-1]
                                tx[k] = float(clean_val)
                            except:
                                tx[k] = 0.0
                        else:
                            tx[k] = str(val).strip()
                    else:
                        tx[k] = ""
                else:
                    tx[k] = ""
                    
            if sheet_name == "GST Transactions":
                tx_type = str(tx.get("type", "")).lower()
                amount = tx.get("total_amount") or 0.0
                if isinstance(amount, str):
                    try:
                        amount = float(amount.replace(",", "").replace("₹", "").strip())
                    except:
                        amount = 0.0
                if "credit" in tx_type:
                    tx["credit"] = amount
                    tx["debit"] = 0.0
                else:
                    tx["debit"] = amount
                    tx["credit"] = 0.0
                    
            for col in ["debit", "credit", "balance"]:
                if tx.get(col) == "" or tx.get(col) is None:
                    tx[col] = 0.0
                else:
                    try:
                        tx[col] = float(tx[col])
                    except:
                        tx[col] = 0.0

            if tx["debit"] > 100000000.0 or tx["credit"] > 100000000.0:
                continue
                
            if tx["debit"] < 0.0 or tx["credit"] < 0.0:
                continue

            if has_val:
                transactions.append(tx)

        seen_txs = set()
        unique_transactions = []
        for tx in transactions:
            tx_key = (tx.get("date"), tx.get("narration"), tx.get("debit"), tx.get("credit"), tx.get("balance"))
            if tx_key not in seen_txs:
                seen_txs.add(tx_key)
                unique_transactions.append(tx)

        # Balance Delta Fallback calculation for statements where Debit/Credit cells are blank but Balance is present
        prev_balance = None
        for tx in unique_transactions:
            d = tx.get("debit", 0.0)
            c = tx.get("credit", 0.0)
            b = tx.get("balance", 0.0)

            if d == 0.0 and c == 0.0 and b > 0.0 and prev_balance is not None and prev_balance > 0.0:
                diff = round(b - prev_balance, 2)
                if diff > 0:
                    tx["credit"] = diff
                    tx["transaction_type"] = "Credit"
                elif diff < 0:
                    tx["debit"] = abs(diff)
                    tx["transaction_type"] = "Debit"

            if b > 0.0:
                prev_balance = b
                
        return unique_transactions

    def load_summary_from_excel(self, excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        meta = {
            "bank_name": "Unknown Bank",
            "account_holder": "Unknown",
            "account_number": "Unknown",
            "period": "Unknown Period",
            "currency": "INR"
        }
        sheet_name = None
        for name in ["Summary", "GST Summary"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if sheet_name:
            ws_sum = wb[sheet_name]
            for row in ws_sum.iter_rows(values_only=True):
                if len(row) >= 2:
                    label = str(row[0]).strip() if row[0] is not None else ""
                    val = row[1]
                    if "Bank Name" in label:
                        meta["bank_name"] = str(val)
                    elif "Account Holder" in label:
                        meta["account_holder"] = str(val)
                    elif "Statement Period" in label:
                        meta["period"] = str(val)
        return meta

    def run_ai_task(self, action_key):
        """Switches the active report view instantly from the pre-generated memory cache."""
        self.active_action_key = action_key

        if not self.active_transactions:
            QMessageBox.warning(self, "No Statement Loaded", "Please upload a statement or select one from history first.")
            return

        # 1. If reports are already prepared, display the view INSTANTLY (0 ms)!
        if self.reports_ready and action_key in self.prepared_reports:
            report_html = GeminiService.clean_html_response(self.prepared_reports[action_key])
            self.current_report_html = report_html
            self.html_wrapper.eval_js("setLoading(false);")
            self.html_wrapper.eval_js(f"setReportHtml({json.dumps(report_html)});")
            
            match = re.search(r"(\d{2,3})\s*/\s*100", report_html)
            if match:
                score_val = match.group(1)
                self.html_wrapper.eval_js(f"document.getElementById('lblScore').innerText = {json.dumps(score_val)};")
            return

        # 2. If still preparing, show non-blocking status
        if not self.reports_ready and hasattr(self, "prepare_thread") and self.prepare_thread is not None and self.prepare_thread.isRunning():
            self.html_wrapper.eval_js("setLoading(true, 'Preparing financial insights...');")
            return

        # 3. Fallback: if not yet prepared, trigger prepare pipeline
        self.start_prepare_all_reports()

    def export_pdf_report(self):
        """Prints the report viewer HTML contents into a PDF file."""
        html_content = self.current_report_html
        if not html_content or "Select a statement" in html_content:
            QMessageBox.warning(self, "No Report Generated", "Please generate an AI report before exporting a PDF.")
            return

        filename = f"AI_Financial_Report_{self.active_metadata.get('bank_name')}_{datetime.datetime.now().strftime('%Y%m%d')}.pdf"
        doc_dir = os.path.expanduser("~/Documents")
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Save AI Financial Report PDF", os.path.join(doc_dir, filename), "PDF Files (*.pdf)"
        )
        
        if not filepath:
            return

        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(filepath)
            printer.setPageMargins(QSize(15, 15), QPrinter.Unit.Millimeter)
            
            doc = QTextDocument()
            doc.setHtml(html_content)
            doc.print_(printer)
            
            Toast.success(self, "✓ PDF Report exported successfully!")
            
            # Auto-create PDF Export Notification
            try:
                from services.notification_service import NotificationService
                user = UserSession.get_current_user()
                user_id = user["id"] if user else "guest"
                NotificationService.create_notification(
                    user_id=user_id,
                    category="parsing_export",
                    title="PDF Export Completed",
                    message=f"PDF report exported successfully: {os.path.basename(filepath)}",
                    action_type="view_report"
                )
                p = self.parent()
                while p:
                    if hasattr(p, "update_notification_badge"):
                        p.update_notification_badge()
                        break
                    p = p.parent()
            except Exception:
                pass

            if os.path.exists(filepath):
                if os.name == 'nt':
                    os.startfile(filepath)
                else:
                    import subprocess
                    subprocess.run(["open", filepath] if os.name == 'posix' else ["xdg-open", filepath])
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Could not export PDF report:\n{e}")

    def open_email_composer(self):
        """Opens Email Composer pre-attaching active AI report."""
        from ui.email_composer_dialog import EmailComposerDialog
        
        attachment = getattr(self, "active_excel_path", None)
        period = getattr(self, "active_metadata", {}).get("period", "")
        bank = getattr(self, "active_metadata", {}).get("bank_name", "")

        dialog = EmailComposerDialog(
            report_type="AI Financial Analysis Report",
            default_attachment=attachment,
            period=period,
            bank_name=bank,
            parent=self
        )
        dialog.exec()
