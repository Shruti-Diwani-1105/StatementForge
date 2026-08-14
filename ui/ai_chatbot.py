import os
import json
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QMessageBox
from PyQt6.QtCore import pyqtSignal, QThread, QTimer

from ui.html_screen_wrapper import HtmlScreenWrapper
from services.gemini_service import GeminiService
from services.history_service import HistoryService
from utils.user_session import UserSession
from settings.toast import Toast

def format_indian_currency(amount: float) -> str:
    """Formats a float value into Indian currency format (e.g., ₹ 15,42,38,560.25)."""
    try:
        is_negative = amount < 0
        amount = abs(amount)
        s = f"{amount:.2f}"
        parts = s.split(".")
        int_part = parts[0]
        dec_part = parts[1]
        
        if len(int_part) <= 3:
            res = int_part
        else:
            thousands_part = int_part[-3:]
            remaining = int_part[:-3]
            pairs = []
            while len(remaining) > 0:
                pairs.append(remaining[-2:])
                remaining = remaining[:-2]
            pairs.reverse()
            res = ",".join(pairs) + "," + thousands_part
            
        formatted = f"₹ {res}.{dec_part}"
        return f"-{formatted}" if is_negative else formatted
    except Exception:
        return f"₹ {amount:,.2f}"


class DBQueryWorker(QThread):
    result_ready = pyqtSignal(object)

    def __init__(self, query_fn, parent=None):
        super().__init__(parent)
        self.query_fn = query_fn

    def run(self):
        try:
            res = self.query_fn()
            self.result_ready.emit(res)
        except Exception as e:
            print(f"DBQueryWorker error: {e}")
            self.result_ready.emit(None)


class AIChatWorker(QThread):
    """
    Background worker thread to execute Gemini chat method without blocking the UI.
    """
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, transactions, chat_history, message, currency="INR"):
        super().__init__()
        self.transactions = transactions
        self.chat_history = chat_history
        self.message = message
        self.currency = currency

    def run(self):
        try:
            if not self.transactions:
                raise ValueError("No statement loaded. Please select a bank statement first.")

            result = GeminiService.chat_with_statement(
                self.transactions, self.chat_history, self.message, self.currency
            )
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class AIChatbotWidget(QWidget):
    """
    Main UI section for AI Chatbot feature powered by HTML + CSS presentation layer.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_theme = "light"
        
        # State Data
        self.active_transactions = []
        self.active_metadata = {
            "bank_name": "Unknown Bank",
            "account_holder": "Unknown",
            "period": "Unknown Period",
            "currency": "INR",
            "total_credit": 0.0,
            "total_debit": 0.0,
            "net_savings": 0.0
        }
        self.chat_history = []
        self.chat_thread = None

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.html_wrapper = HtmlScreenWrapper("web/ai_chatbot.html", self)
        layout.addWidget(self.html_wrapper)

        # Connect document title / WebBridge IPC commands
        self.html_wrapper.web_view.titleChanged.connect(self.handle_web_commands)
        
        # Initial dropdown load timer
        QTimer.singleShot(600, self.load_history_dropdown)

    def update_theme_style(self, theme: str = "light"):
        """Updates HTML UI theme styling ('light' or 'dark')."""
        self.current_theme = theme.lower().strip() if isinstance(theme, str) else "light"
        if self.current_theme == "dark":
            self.html_wrapper.eval_js("document.body.classList.add('dark-mode');")
        else:
            self.html_wrapper.eval_js("document.body.classList.remove('dark-mode');")

    def handle_web_commands(self, title: str):
        """Dispatches commands sent from JavaScript UI via document.title IPC."""
        if not title or not title.startswith("app-cmd:"):
            return
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, lambda: self._process_web_commands(title))

    def _process_web_commands(self, title: str):
        parts = title.split(":", 2)
        cmd = parts[1] if len(parts) > 1 else ""
        raw_payload = parts[2] if len(parts) > 2 else ""

        if cmd == "auditor_refresh_history":
            self.load_history_dropdown()
        elif cmd == "auditor_select_statement":
            try:
                data = json.loads(raw_payload)
                excel_path = data.get("path")
                if excel_path:
                    self.on_statement_selected_by_path(excel_path)
            except Exception as e:
                print(f"select error: {e}")
        elif cmd == "auditor_send_chat":
            try:
                data = json.loads(raw_payload)
                msg = data.get("message", "")
                if msg:
                    self.send_chat_message(msg)
            except Exception:
                pass

    def set_active_statement(self, payload):
        """Allows direct programmatic insertion of parsed statements."""
        if not payload or not payload.get("transactions"):
            return

        self.active_transactions = payload["transactions"]
        
        self.active_metadata = {
            "bank_name": payload.get("bank_name", "Unknown Bank"),
            "account_holder": payload.get("account_holder", "Unknown"),
            "period": payload.get("period", "Unknown Period"),
            "account_number": payload.get("account_number", "Unknown"),
            "currency": payload.get("currency", "INR")
        }
        
        credits = 0.0
        debits = 0.0
        for tx in self.active_transactions:
            credits += tx.get("credit") or 0.0
            debits += tx.get("debit") or 0.0
            
        self.active_metadata["total_credit"] = credits
        self.active_metadata["total_debit"] = debits
        self.active_metadata["net_savings"] = credits - debits

        # Reset chat history
        self.chat_history = []
        chat_reset_js = """
        var container = document.getElementById('chatContainer');
        if (container) {
            container.innerHTML = `
                <div class="chat-bubble advisor-bubble">
                    <div class="bubble-avatar">🤖</div>
                    <div class="bubble-content">
                        <div class="bubble-header">
                            <span class="sender-name">AI Chatbot</span>
                            <span class="bubble-time">Now</span>
                        </div>
                        <div class="bubble-text">
                            Hello! I am your AI Chatbot. Ask me any questions about this statement like:
                            <br>• What are my top expenses?
                            <br>• Are there any duplicate UPI transactions?
                            <br>• What subscriptions did I pay for?
                            <br>• Where am I spending the most?
                        </div>
                    </div>
                </div>
            `;
        }
        """
        self.html_wrapper.eval_js(chat_reset_js)

        # Update metrics preview in HTML
        self.update_metrics_ui()

    def update_metrics_ui(self):
        """Fills UI metric values from active statement details into HTML."""
        bank = self.active_metadata.get("bank_name", "-")
        period = self.active_metadata.get("period", "-")
        
        credits_str = format_indian_currency(self.active_metadata.get("total_credit", 0.0))
        debits_str = format_indian_currency(self.active_metadata.get("total_debit", 0.0))
        net = self.active_metadata.get("net_savings", 0.0)
        savings_str = format_indian_currency(net)
        is_positive = net >= 0
        
        score = 92
        if len(self.active_transactions) > 0:
            score = min(98, max(60, 100 - int(len(self.active_transactions) * 0.15)))
            score_str = f"{score}%"
        else:
            score_str = "-"
            
        js_code = (
            f"setMetrics("
            f"{json.dumps(bank)}, "
            f"{json.dumps(period)}, "
            f"{json.dumps(credits_str)}, "
            f"{json.dumps(debits_str)}, "
            f"{json.dumps(savings_str)}, "
            f"{json.dumps(score_str)}, "
            f"{'true' if is_positive else 'false'});"
        )
        self.html_wrapper.eval_js(js_code)

    def _safe_run_query(self, query_fn, callback_fn):
        worker = DBQueryWorker(query_fn, self)
        if not hasattr(self, "_active_workers"):
            self._active_workers = []
        self._active_workers.append(worker)
        
        def handle_result(res):
            if worker in self._active_workers:
                self._active_workers.remove(worker)
            if res is not None:
                callback_fn(res)
                
        worker.result_ready.connect(handle_result)
        worker.start()

    def load_history_dropdown(self):
        """Queries local history database logs and updates dropdown list in HTML."""
        user = UserSession.get_current_user()
        user_id = user["id"] if user else "guest"
        
        def db_query():
            return HistoryService.get_history_logs(user_id=user_id)
            
        def db_callback(logs):
            completed_logs = [log for log in logs if log.get("status") == "Completed" and log.get("excel_path")]
            
            options_js = "(function(){ var cb = document.getElementById('statementCb'); if(!cb) return; cb.innerHTML = '';"
            if not completed_logs:
                options_js += "cb.innerHTML += `<option value=''>No parsed statements found in history.</option>`;"
            else:
                options_js += "cb.innerHTML += `<option value=''>Select from parsed statement history...</option>`;"
                for log in completed_logs:
                    pdf_path = log.get("pdf_path", "")
                    excel_path = log.get("excel_path", "")
                    filename = os.path.basename(pdf_path) if pdf_path else "Statement.pdf"
                    upload_date = log.get("upload_date")
                    if hasattr(upload_date, "strftime"):
                        date_str = upload_date.strftime("%Y-%m-%d")
                    elif isinstance(upload_date, str):
                        date_str = upload_date[:10]
                    else:
                        date_str = str(upload_date or "")[:10]
                    bank = log.get("bank_name", "Unknown Bank")
                    display_text = f"{bank} ({date_str}) - {filename}"
                    escaped_path = json.dumps(excel_path)
                    options_js += f"cb.innerHTML += `<option value={escaped_path}>{display_text}</option>`;"
            options_js += "})();"
            self.html_wrapper.eval_js(options_js)

        self._safe_run_query(db_query, db_callback)

    def on_statement_selected_by_path(self, excel_path):
        """Loads and processes transaction details from history Excel path when selected in HTML."""
        if not excel_path:
            return
            
        if "_GST_Report.xlsx" in excel_path:
            possible_std = excel_path.replace("_GST_Report.xlsx", ".xlsx")
            if os.path.exists(possible_std):
                excel_path = possible_std
                
        if not os.path.exists(excel_path):
            return
            
        try:
            self.active_excel_path = excel_path
            transactions = self.load_transactions_from_excel(excel_path)
            meta = self.load_summary_from_excel(excel_path)
            
            payload = {
                "transactions": transactions,
                "bank_name": meta.get("bank_name", "Unknown Bank"),
                "account_holder": meta.get("account_holder", "Unknown"),
                "account_number": meta.get("account_number", "Unknown"),
                "period": meta.get("period", "Unknown Period"),
                "currency": meta.get("currency", "INR")
            }
            
            self.set_active_statement(payload)
            Toast.success(self, "✓ Statement loaded for Chatbot!")
        except Exception as e:
            QMessageBox.critical(self, "Error Loading Excel", f"Could not load transaction sheets from Excel archive:\n{e}")
            self.load_history_dropdown()

    def load_transactions_from_excel(self, excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_name = None
        for name in ["Transactions", "GST Transactions"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            raise ValueError("Spreadsheet does not contain a transaction ledger sheet.")
            
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
            
        headers = rows[0]
        data_rows = rows[1:]
        
        col_mapping = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            h_lower = str(header).lower()
            if "date" in h_lower and "value" not in h_lower:
                col_mapping["date"] = idx
            elif "description" in h_lower or "narration" in h_lower or "particulars" in h_lower:
                col_mapping["narration"] = idx
            elif "debit" in h_lower:
                col_mapping["debit"] = idx
            elif "credit" in h_lower:
                col_mapping["credit"] = idx
            elif "balance" in h_lower:
                col_mapping["balance"] = idx
            elif "type" in h_lower:
                col_mapping["type"] = idx
            elif "total amount" in h_lower or "amount" in h_lower:
                col_mapping["total_amount"] = idx
                
        transactions = []
        for r in data_rows:
            tx = {}
            has_val = False
            for k in ["date", "narration", "debit", "credit", "balance", "type", "total_amount"]:
                idx = col_mapping.get(k)
                if idx is not None and idx < len(r):
                    val = r[idx]
                    if val is not None:
                        has_val = True
                        if k == "date" and hasattr(val, "strftime"):
                            tx[k] = val.strftime("%Y-%m-%d")
                        elif k in ["debit", "credit", "balance", "total_amount"]:
                            try:
                                clean_val = str(val).replace(",", "").replace("₹", "").strip()
                                if clean_val.endswith("-"):
                                    clean_val = "-" + clean_val[:-1]
                                tx[k] = float(clean_val)
                            except:
                                tx[k] = 0.0
                        else:
                            tx[k] = str(val).strip()
                    else:
                        tx[k] = ""
                else:
                    tx[k] = ""
                    
            if sheet_name == "GST Transactions":
                tx_type = str(tx.get("type", "")).lower()
                amount = tx.get("total_amount") or 0.0
                if isinstance(amount, str):
                    try:
                        amount = float(amount.replace(",", "").replace("₹", "").strip())
                    except:
                        amount = 0.0
                if "credit" in tx_type:
                    tx["credit"] = amount
                    tx["debit"] = 0.0
                else:
                    tx["debit"] = amount
                    tx["credit"] = 0.0
                    
            for col in ["debit", "credit", "balance"]:
                if tx.get(col) == "" or tx.get(col) is None:
                    tx[col] = 0.0
                else:
                    try:
                        tx[col] = float(tx[col])
                    except:
                        tx[col] = 0.0

            if tx["debit"] > 100000000.0 or tx["credit"] > 100000000.0:
                continue
                
            if tx["debit"] < 0.0 or tx["credit"] < 0.0:
                continue

            if has_val:
                transactions.append(tx)

        seen_txs = set()
        unique_transactions = []
        for tx in transactions:
            tx_key = (tx.get("date"), tx.get("narration"), tx.get("debit"), tx.get("credit"), tx.get("balance"))
            if tx_key not in seen_txs:
                seen_txs.add(tx_key)
                unique_transactions.append(tx)
                
        return unique_transactions

    def load_summary_from_excel(self, excel_path):
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        meta = {
            "bank_name": "Unknown Bank",
            "account_holder": "Unknown",
            "account_number": "Unknown",
            "period": "Unknown Period",
            "currency": "INR"
        }
        sheet_name = None
        for name in ["Summary", "GST Summary"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        if sheet_name:
            ws_sum = wb[sheet_name]
            for row in ws_sum.iter_rows(values_only=True):
                if len(row) >= 2:
                    label = str(row[0]).strip() if row[0] is not None else ""
                    val = row[1]
                    if "Bank Name" in label:
                        meta["bank_name"] = str(val)
                    elif "Account Holder" in label:
                        meta["account_holder"] = str(val)
                    elif "Statement Period" in label:
                        meta["period"] = str(val)
        return meta

    def send_chat_message(self, msg):
        """Sends user text message contextually to Gemini alongside current transaction logs."""
        if not msg:
            return

        if not self.active_transactions:
            QMessageBox.warning(self, "No Statement Loaded", "Please upload a statement or select one from history first.")
            return

        self.chat_history.append({"role": "user", "content": msg})
        self.html_wrapper.eval_js("showThinkingIndicator();")

        # Trigger background chat thread
        self.chat_thread = AIChatWorker(
            self.active_transactions,
            self.chat_history,
            msg,
            currency=self.active_metadata.get("currency", "INR")
        )

        def on_finished(reply):
            self.html_wrapper.eval_js("removeThinkingIndicator();")
            self.chat_thread.deleteLater()
            self.chat_thread = None

            self.chat_history.append({"role": "assistant", "content": reply})
            self.html_wrapper.eval_js(f"appendChatBubble('advisor', {json.dumps(reply)});")

        def on_error(err):
            self.html_wrapper.eval_js("removeThinkingIndicator();")
            self.chat_thread.deleteLater()
            self.chat_thread = None
            
            err_text = f"⚠️ Failed to get answer: {err}"
            self.html_wrapper.eval_js(f"appendChatBubble('advisor', {json.dumps(err_text)});")

        self.chat_thread.finished.connect(on_finished)
        self.chat_thread.error.connect(on_error)
        self.chat_thread.start()
