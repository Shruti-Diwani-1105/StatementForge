import os
import datetime
from xml.sax.saxutils import escape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TallyService:
    """
    Handles business logic for Tally XML Voucher Import and Tally Excel Template export.
    Includes direct sync via Tally HTTP API and Master Ledger auto-creation.
    """

    @classmethod
    def parse_to_tally_date(cls, date_str):
        """Parses various date formats and returns Tally-formatted YYYYMMDD string."""
        if not date_str:
            return datetime.datetime.now().strftime("%Y%m%d")
        
        date_str = str(date_str).strip()
        # Clean any timestamp parts
        if " " in date_str:
            date_str = date_str.split(" ")[0]
            
        date_formats = [
            "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y",
            "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d %H:%M:%S",
            "%d %b %Y", "%d-%b-%Y", "%d-%b-%y"
        ]
        
        for fmt in date_formats:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y%m%d")
            except ValueError:
                pass
                
        # Return clean digits if already YYYYMMDD
        digits = "".join(filter(str.isdigit, date_str))
        if len(digits) == 8:
            return digits
            
        return datetime.datetime.now().strftime("%Y%m%d")

    @classmethod
    def auto_derive_voucher_type(cls, tx, narration=""):
        """Auto-categorizes a transaction into Payment, Receipt, or Contra voucher types."""
        # Check if cash related
        narr_upper = narration.upper()
        if any(kw in narr_upper for kw in ("CASH", "ATM", "SELF", "CASH WITHDRAWAL", "CASH DEPOSIT")):
            return "Contra"
            
        # Check debit vs credit
        debit_amt = float(tx.get("debit", 0.0) or 0.0)
        credit_amt = float(tx.get("credit", 0.0) or 0.0)
        tx_type = str(tx.get("type", "Debit")).lower()
        
        if debit_amt > 0 or "debit" in tx_type or "withdrawal" in tx_type:
            return "Payment"
        else:
            return "Receipt"

    @classmethod
    def generate_ledger_masters_xml(cls, ledger_mappings):
        """
        Generates XML to create multiple ledger accounts in Tally under specified groups.
        ledger_mappings: List of dicts, e.g. [{"name": "HDFC Bank", "parent": "Bank Accounts"}]
        """
        xml_str = """<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <IMPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>All Masters</REPORTNAME>
        <STATICVARIABLES>
          <SVCURRENTCOMPANY>StatementForge Company</SVCURRENTCOMPANY>
        </STATICVARIABLES>
      </REQUESTDESC>
      <REQUESTDATA>
"""
        for mapping in ledger_mappings:
            name = mapping["name"]
            parent = mapping["parent"]
            xml_str += f"""        <TALLYMESSAGE xmlns:UDF="TallyUDF">
          <LEDGER NAME="{escape(name)}" ACTION="Create">
            <NAME>{escape(name)}</NAME>
            <PARENT>{escape(parent)}</PARENT>
          </LEDGER>
        </TALLYMESSAGE>
"""
        xml_str += """      </REQUESTDATA>
    </IMPORTDATA>
  </BODY>
</ENVELOPE>"""
        return xml_str

    @classmethod
    def generate_tally_xml(cls, transactions, bank_ledger_name, default_counter_ledger="Suspense Account"):
        """
        Generates a standard Tally XML voucher import string.
        
        Sign convention rules in Tally XML:
        - Debit entry: <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE> and negative <AMOUNT> (e.g. -500.00)
        - Credit entry: <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE> and positive <AMOUNT> (e.g. 500.00)
        """
        xml_str = """<ENVELOPE>
  <HEADER>
    <TALLYREQUEST>Import Data</TALLYREQUEST>
  </HEADER>
  <BODY>
    <IMPORTDATA>
      <REQUESTDESC>
        <REPORTNAME>Vouchers</REPORTNAME>
        <STATICVARIABLES>
          <SVCURRENTCOMPANY>StatementForge Company</SVCURRENTCOMPANY>
        </STATICVARIABLES>
      </REQUESTDESC>
      <REQUESTDATA>
"""
        for tx in transactions:
            date_str = tx.get("date", "")
            tally_date = cls.parse_to_tally_date(date_str)
            
            narration = tx.get("narration", "Bank Transaction")
            debit_amt = float(tx.get("debit", 0.0) or 0.0)
            credit_amt = float(tx.get("credit", 0.0) or 0.0)
            amount = debit_amt if debit_amt > 0 else credit_amt
            if amount <= 0:
                amount = float(tx.get("amount", 0.0) or tx.get("total_amount", 0.0) or 0.0)
                amount = abs(amount)

            # Determine Voucher Type & Ledgers
            vch_type = tx.get("voucher_type")
            if not vch_type:
                vch_type = cls.auto_derive_voucher_type(tx, narration)
                
            row_ledger = tx.get("ledger_name") or default_counter_ledger
            
            is_withdrawal = (debit_amt > 0 or "debit" in str(tx.get("type", "Debit")).lower() or "withdrawal" in str(tx.get("type", "Debit")).lower())
            
            if vch_type == "Payment":
                dr_ledger = row_ledger
                cr_ledger = bank_ledger_name
            elif vch_type == "Receipt":
                dr_ledger = bank_ledger_name
                cr_ledger = row_ledger
            elif vch_type == "Contra":
                if is_withdrawal:
                    dr_ledger = "Cash"
                    cr_ledger = bank_ledger_name
                else:
                    dr_ledger = bank_ledger_name
                    cr_ledger = "Cash"
            else:  # Journal or custom
                if is_withdrawal:
                    dr_ledger = row_ledger
                    cr_ledger = bank_ledger_name
                else:
                    dr_ledger = bank_ledger_name
                    cr_ledger = row_ledger

            # Format vouchers xml
            xml_str += f"""        <TALLYMESSAGE xmlns:UDF="TallyUDF">
          <VOUCHER VCHTYPE="{escape(vch_type)}" ACTION="Create" OBJVIEW="Accounting Voucher View">
            <DATE>{tally_date}</DATE>
            <VOUCHERTYPENAME>{escape(vch_type)}</VOUCHERTYPENAME>
            <PARTYLEDGERNAME>{escape(dr_ledger)}</PARTYLEDGERNAME>
            <EFFECTIVEDATE>{tally_date}</EFFECTIVEDATE>
            <NARRATION>{escape(narration)}</NARRATION>
            <ALLLEDGERENTRIES.LIST>
              <LEDGERNAME>{escape(dr_ledger)}</LEDGERNAME>
              <ISDEEMEDPOSITIVE>Yes</ISDEEMEDPOSITIVE>
              <AMOUNT>-{amount:.2f}</AMOUNT>
            </ALLLEDGERENTRIES.LIST>
            <ALLLEDGERENTRIES.LIST>
              <LEDGERNAME>{escape(cr_ledger)}</LEDGERNAME>
              <ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>
              <AMOUNT>{amount:.2f}</AMOUNT>
            </ALLLEDGERENTRIES.LIST>
          </VOUCHER>
        </TALLYMESSAGE>
"""
        xml_str += """      </REQUESTDATA>
    </IMPORTDATA>
  </BODY>
</ENVELOPE>"""
        return xml_str

    @classmethod
    def generate_tally_excel(cls, transactions, bank_ledger_name, default_counter_ledger="Suspense Account", output_path=None):
        """
        Generates a flat Tally Prime-friendly Excel voucher template.
        """
        if not output_path:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Tally Vouchers"
        
        ws.views.sheetView[0].showGridLines = True
        
        headers = ["Date", "Voucher Type", "Voucher No", "Dr Ledger", "Cr Ledger", "Amount", "Narration"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        row_font = Font(name="Segoe UI", size=10)
        
        for tx in transactions:
            date_str = tx.get("date", "")
            try:
                t_date = cls.parse_to_tally_date(date_str)
                dt = datetime.datetime.strptime(t_date, "%Y%m%d")
                formatted_date = dt.strftime("%d-%m-%Y")
            except:
                formatted_date = date_str

            narration = tx.get("narration", "Bank Transaction")
            debit_amt = float(tx.get("debit", 0.0) or 0.0)
            credit_amt = float(tx.get("credit", 0.0) or 0.0)
            amount = debit_amt if debit_amt > 0 else credit_amt
            if amount <= 0:
                amount = float(tx.get("amount", 0.0) or tx.get("total_amount", 0.0) or 0.0)
                amount = abs(amount)

            vch_type = tx.get("voucher_type")
            if not vch_type:
                vch_type = cls.auto_derive_voucher_type(tx, narration)
                
            row_ledger = tx.get("ledger_name") or default_counter_ledger
            
            is_withdrawal = (debit_amt > 0 or "debit" in str(tx.get("type", "Debit")).lower() or "withdrawal" in str(tx.get("type", "Debit")).lower())
            
            if vch_type == "Payment":
                dr_ledger = row_ledger
                cr_ledger = bank_ledger_name
            elif vch_type == "Receipt":
                dr_ledger = bank_ledger_name
                cr_ledger = row_ledger
            elif vch_type == "Contra":
                if is_withdrawal:
                    dr_ledger = "Cash"
                    cr_ledger = bank_ledger_name
                else:
                    dr_ledger = bank_ledger_name
                    cr_ledger = "Cash"
            else:
                if is_withdrawal:
                    dr_ledger = row_ledger
                    cr_ledger = bank_ledger_name
                else:
                    dr_ledger = bank_ledger_name
                    cr_ledger = "Cash"

            ref_no = tx.get("ref_no", "") or tx.get("cheque_no", "") or ""
            
            row_data = [formatted_date, vch_type, ref_no, dr_ledger, cr_ledger, amount, narration]
            ws.append(row_data)
            
        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = row_font
                cell.border = thin_border
                
                if c_idx in (1, 2, 3):
                    cell.alignment = center_align
                elif c_idx in (4, 5, 7):
                    cell.alignment = left_align
                elif c_idx == 6:
                    cell.alignment = right_align
                    cell.number_format = '0.00'

        ws.row_dimensions[1].height = 28
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(output_path)
        return output_path

    @classmethod
    def push_to_tally_server(cls, url, xml_data):
        """
        Sends an XML POST request to Tally acts-as-HTTP local Server.
        Returns a dict with success indicator, created counts, and log details.
        """
        import urllib.request
        import urllib.error
        import xml.etree.ElementTree as ET
        
        headers = {
            'Content-Type': 'text/xml; charset=utf-8',
        }
        
        req = urllib.request.Request(
            url, 
            data=xml_data.encode('utf-8'), 
            headers=headers, 
            method='POST'
        )
        
        try:
            with urllib.request.urlopen(req, timeout=5) as response:
                resp_bytes = response.read()
                resp_str = resp_bytes.decode('utf-8')
                
                created = 0
                errors = 0
                message = "Success"
                success = True
                
                try:
                    root = ET.fromstring(resp_str)
                    
                    created_el = root.find(".//CREATED")
                    errors_el = root.find(".//ERRORS")
                    line_err_el = root.find(".//LINEERROR")
                    
                    if created_el is not None:
                        created = int(created_el.text or 0)
                    if errors_el is not None:
                        errors = int(errors_el.text or 0)
                        if errors > 0 and line_err_el is not None:
                            message = line_err_el.text or "Error importing vouchers."
                            success = False
                except Exception as parse_ex:
                    if "<CREATED>" in resp_str:
                        import re
                        m = re.search(r"<CREATED>(\d+)</CREATED>", resp_str)
                        if m: created = int(m.group(1))
                        m_err = re.search(r"<ERRORS>(\d+)</ERRORS>", resp_str)
                        if m_err: errors = int(m_err.group(1))
                        if errors > 0:
                            success = False
                            message = "Tally validation errors found."
                    else:
                        message = f"Response parsed with warning: {parse_ex}"
                
                return {
                    "success": success and errors == 0,
                    "created": created,
                    "errors": errors,
                    "raw_response": resp_str,
                    "message": message
                }
                
        except urllib.error.URLError as e:
            return {
                "success": False,
                "created": 0,
                "errors": 1,
                "raw_response": "",
                "message": f"Connection to Tally failed: {e.reason}. Make sure Tally acts as an HTTP server on port 9000."
            }
        except Exception as e:
            return {
                "success": False,
                "created": 0,
                "errors": 1,
                "raw_response": "",
                "message": f"Sync transmission error: {str(e)}"
            }
