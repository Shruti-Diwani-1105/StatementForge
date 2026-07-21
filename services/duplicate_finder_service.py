import os
import re
import datetime
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DuplicateFinderService:
    """
    Core engine for detecting, clustering, auto-resolving, and exporting
    duplicate bank statement transactions across single or multiple files.
    """

    @staticmethod
    def _clean_str(val):
        if val is None:
            return ""
        return str(val).strip()

    @staticmethod
    def _normalize_text(text):
        """Normalizes text for fuzzy matching by lowercasing and stripping non-alphanumeric noise."""
        if not text:
            return ""
        text = text.lower()
        # Remove common prefixes/suffixes like Ref No, UPI/, etc.
        text = re.sub(r'[^a-z0-9\s]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    @staticmethod
    def _parse_amount(val):
        """Helper to parse amount floats safely."""
        if val is None:
            return 0.0
        s = str(val).replace(",", "").strip()
        if not s:
            return 0.0
        try:
            return abs(float(s))
        except ValueError:
            return 0.0

    @staticmethod
    def _parse_date(date_str):
        """Helper to convert string dates to datetime.date objects for comparison."""
        if not date_str:
            return None
        s = str(date_str).strip()
        # Common formats
        formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y",
            "%d/%m/%y", "%d-%m-%y", "%m/%d/%Y"
        ]
        for fmt in formats:
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    @classmethod
    def analyze_statement(cls, transactions, options=None):
        """
        Analyzes a list of transaction dictionaries and returns duplicate clusters and statistics.

        :param transactions: List of dicts with keys (date, narration, debit, credit, balance, ref_no, etc.)
        :param options: Dict with detection flags:
            - exact_match (bool): Check 100% exact row duplicates (default True)
            - potential_match (bool): Check fuzzy/narration + amount matches (default True)
            - date_window_days (int): Tolerance for date differences in days (default 2)
            - similarity_threshold (float): Narration similarity (0.0 to 1.0, default 0.75)
        :return: Dict containing 'clusters', 'stats', and 'annotated_transactions'
        """
        if options is None:
            options = {}

        exact_match_opt = options.get("exact_match", True)
        potential_match_opt = options.get("potential_match", True)
        date_window_days = int(options.get("date_window_days", 2))
        similarity_thresh = float(options.get("similarity_threshold", 0.75))

        if not transactions:
            return {
                "clusters": [],
                "stats": {
                    "total_transactions": 0,
                    "duplicate_clusters": 0,
                    "duplicate_entries": 0,
                    "flagged_debit_sum": 0.0,
                    "flagged_credit_sum": 0.0,
                    "cleanliness_score": 100.0
                },
                "annotated_transactions": []
            }

        # Prepare normalized items
        items = []
        for idx, tx in enumerate(transactions):
            date_obj = cls._parse_date(tx.get("date", ""))
            debit_amt = cls._parse_amount(tx.get("debit", ""))
            credit_amt = cls._parse_amount(tx.get("credit", ""))
            net_amt = debit_amt if debit_amt > 0 else credit_amt
            tx_type = "Debit" if debit_amt > 0 else ("Credit" if credit_amt > 0 else "Neutral")
            
            raw_narr = cls._clean_str(tx.get("narration", ""))
            norm_narr = cls._normalize_text(raw_narr)
            ref_no = cls._clean_str(tx.get("ref_no", ""))
            source_file = cls._clean_str(tx.get("source_file", "Current Statement"))
            
            items.append({
                "id": f"tx_{idx}",
                "original_index": idx,
                "raw": tx,
                "date_str": cls._clean_str(tx.get("date", "")),
                "date_obj": date_obj,
                "raw_narration": raw_narr,
                "norm_narration": norm_narr,
                "debit": debit_amt,
                "credit": credit_amt,
                "amount": net_amt,
                "type": tx_type,
                "ref_no": ref_no,
                "source_file": source_file,
                "balance": cls._clean_str(tx.get("balance", "")),
                "cluster_id": None,
                "match_type": None
            })

        visited = set()
        clusters = []
        cluster_counter = 1

        # 1. Exact Match Pass
        if exact_match_opt:
            exact_groups = {}
            for item in items:
                # Key for exact match
                key = (
                    item["date_str"],
                    item["raw_narration"].lower(),
                    item["debit"],
                    item["credit"],
                    item["balance"]
                )
                exact_groups.setdefault(key, []).append(item)

            for key, group in exact_groups.items():
                if len(group) > 1:
                    c_id = f"CLUSTER_EXACT_{cluster_counter}"
                    cluster_counter += 1
                    for it in group:
                        visited.add(it["id"])
                        it["cluster_id"] = c_id
                        it["match_type"] = "Exact Duplicate"

                    debit_sum = sum(it["debit"] for it in group[1:]) # Count excess as flagged
                    credit_sum = sum(it["credit"] for it in group[1:])
                    
                    clusters.append({
                        "id": c_id,
                        "title": f"Exact Duplicate Cluster #{cluster_counter-1}",
                        "match_type": "Exact Match",
                        "confidence": 100,
                        "risk_level": "High",
                        "badge_color": "#EF4444",
                        "badge_bg": "#FEF2F2",
                        "reason": f"{len(group)} transactions share identical Date ({group[0]['date_str']}), Narration, and Amount ({group[0]['amount']:,.2f}).",
                        "items": group,
                        "flagged_debit": debit_sum,
                        "flagged_credit": credit_sum
                    })

        # 2. Potential / Fuzzy Match Pass
        if potential_match_opt:
            n_items = len(items)
            for i in range(n_items):
                item_i = items[i]
                if item_i["id"] in visited or item_i["amount"] == 0:
                    continue

                current_cluster = [item_i]
                match_reasons = []

                for j in range(i + 1, n_items):
                    item_j = items[j]
                    if item_j["id"] in visited or item_j["amount"] == 0:
                        continue

                    # Amount must be equal (or within 0.01 tolerance)
                    if abs(item_i["amount"] - item_j["amount"]) > 0.01:
                        continue

                    # Transaction type (debit vs credit) must match
                    if item_i["type"] != item_j["type"]:
                        continue

                    # Check Date Window
                    date_diff = None
                    if item_i["date_obj"] and item_j["date_obj"]:
                        date_diff = abs((item_i["date_obj"] - item_j["date_obj"]).days)

                    date_match = (date_diff is not None and date_diff <= date_window_days) or (item_i["date_str"] == item_j["date_str"])

                    if not date_match:
                        continue

                    # Check Narration Similarity
                    sim = SequenceMatcher(None, item_i["norm_narration"], item_j["norm_narration"]).ratio()
                    
                    # Also check cross-statement flag
                    is_cross_file = (item_i["source_file"] != item_j["source_file"])

                    if sim >= similarity_thresh or is_cross_file:
                        current_cluster.append(item_j)
                        if is_cross_file:
                            match_reasons.append(f"Identical amount found across multiple files ('{item_i['source_file']}' & '{item_j['source_file']}')")
                        elif date_diff and date_diff > 0:
                            match_reasons.append(f"Matching amount ({item_i['amount']:,.2f}) within {date_diff} day(s)")
                        else:
                            match_reasons.append(f"Fuzzy narration match ({int(sim*100)}% similarity)")

                if len(current_cluster) > 1:
                    is_cross = any(it["source_file"] != current_cluster[0]["source_file"] for it in current_cluster)
                    m_type = "Cross-Statement Duplicate" if is_cross else "Potential Duplicate"
                    c_id = f"CLUSTER_POTENTIAL_{cluster_counter}"
                    cluster_counter += 1

                    for it in current_cluster:
                        visited.add(it["id"])
                        it["cluster_id"] = c_id
                        it["match_type"] = m_type

                    debit_sum = sum(it["debit"] for it in current_cluster[1:])
                    credit_sum = sum(it["credit"] for it in current_cluster[1:])

                    confidence = 90 if m_type == "Cross-Statement Duplicate" else 75
                    badge_col = "#2563EB" if is_cross else "#F59E0B"
                    badge_bg = "#EFF6FF" if is_cross else "#FFFBEB"

                    clusters.append({
                        "id": c_id,
                        "title": f"{m_type} Cluster #{cluster_counter-1}",
                        "match_type": m_type,
                        "confidence": confidence,
                        "risk_level": "Medium",
                        "badge_color": badge_col,
                        "badge_bg": badge_bg,
                        "reason": match_reasons[0] if match_reasons else f"Flagged potential double charge of {current_cluster[0]['amount']:,.2f}.",
                        "items": current_cluster,
                        "flagged_debit": debit_sum,
                        "flagged_credit": credit_sum
                    })

        # Calculate Overall Stats
        total_tx = len(transactions)
        total_clusters = len(clusters)
        dup_entries_count = sum(len(c["items"]) for c in clusters)
        excess_entries_count = sum(len(c["items"]) - 1 for c in clusters)
        
        total_flagged_debit = sum(c["flagged_debit"] for c in clusters)
        total_flagged_credit = sum(c["flagged_credit"] for c in clusters)

        cleanliness = 100.0 if total_tx == 0 else max(0.0, round(((total_tx - excess_entries_count) / total_tx) * 100, 1))

        stats = {
            "total_transactions": total_tx,
            "duplicate_clusters": total_clusters,
            "duplicate_entries": dup_entries_count,
            "excess_entries": excess_entries_count,
            "flagged_debit_sum": total_flagged_debit,
            "flagged_credit_sum": total_flagged_credit,
            "cleanliness_score": cleanliness
        }

        return {
            "clusters": clusters,
            "stats": stats,
            "annotated_transactions": items
        }

    @classmethod
    def analyze_multiple_statements(cls, statements_payload_list, options=None):
        """
        Combines transactions from multiple statement payloads and flags cross-statement duplicates.
        """
        combined_txs = []
        for p in statements_payload_list:
            file_name = p.get("file_name", p.get("bank_name", "Statement"))
            txs = p.get("transactions", [])
            for tx in txs:
                tx_copy = dict(tx)
                tx_copy["source_file"] = file_name
                combined_txs.append(tx_copy)

        return cls.analyze_statement(combined_txs, options)

    @classmethod
    def apply_auto_resolution(cls, clusters, strategy="keep_first"):
        """
        Applies an automated resolution rule across all duplicate clusters.

        :param clusters: List of cluster objects returned by analyze_statement
        :param strategy: 'keep_first', 'keep_last', or 'keep_highest_ref'
        :return: Dict mapping item_id -> action ('keep' or 'remove')
        """
        decisions = {}
        for c in clusters:
            items = c["items"]
            if not items:
                continue

            if strategy == "keep_last":
                keep_item = items[-1]
            elif strategy == "keep_highest_ref":
                keep_item = max(items, key=lambda x: len(x.get("ref_no", "")))
            else: # default keep_first
                keep_item = items[0]

            for it in items:
                if it["id"] == keep_item["id"]:
                    decisions[it["id"]] = "keep"
                else:
                    decisions[it["id"]] = "remove"

        return decisions

    @classmethod
    def export_duplicate_report(cls, clusters, stats, output_path):
        """
        Generates a professional openpyxl Excel audit report with Summary, Duplicate Clusters, and Cleaned Ledger.
        """
        wb = openpyxl.Workbook()

        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_exact = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
        fill_potential = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        fill_cross = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # ----------------------------------------------------
        # SHEET 1: AUDIT SUMMARY
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Audit Summary"
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.merge_cells("A1:E1")
        title_cell = ws_sum["A1"]
        title_cell.value = "StatementForge - Duplicate Transaction Audit Report"
        title_cell.font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws_sum["A2"] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_sum["A2"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

        # KPI Metrics Table
        ws_sum["A4"] = "Audit Key Metrics"
        ws_sum["A4"].font = Font(name="Calibri", size=12, bold=True, color="1E293B")

        kpis = [
            ("Total Transactions Scanned", stats.get("total_transactions", 0)),
            ("Duplicate Clusters Identified", stats.get("duplicate_clusters", 0)),
            ("Excess Duplicate Entries", stats.get("excess_entries", 0)),
            ("Flagged Duplicate Debits", f"₹ {stats.get('flagged_debit_sum', 0.0):,.2f}"),
            ("Flagged Duplicate Credits", f"₹ {stats.get('flagged_credit_sum', 0.0):,.2f}"),
            ("Statement Cleanliness Score", f"{stats.get('cleanliness_score', 100.0)}%")
        ]

        row = 5
        for metric, val in kpis:
            ws_sum.cell(row=row, column=1, value=metric).font = Font(bold=True, color="475569")
            c_val = ws_sum.cell(row=row, column=2, value=val)
            c_val.font = Font(bold=True, color="0F172A")
            ws_sum.cell(row=row, column=1).border = thin_border
            c_val.border = thin_border
            row += 1

        # ----------------------------------------------------
        # SHEET 2: DUPLICATE CLUSTERS
        # ----------------------------------------------------
        ws_cls = wb.create_sheet(title="Flagged Duplicate Clusters")
        ws_cls.views.sheetView[0].showGridLines = True

        headers = ["Cluster ID", "Match Type", "Confidence", "Source File", "Date", "Narration", "Debit (₹)", "Credit (₹)", "Balance (₹)", "Ref No", "Audit Reason"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_cls.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        r = 2
        for c in clusters:
            m_type = c["match_type"]
            fill_bg = fill_exact if m_type == "Exact Match" else (fill_cross if m_type == "Cross-Statement Duplicate" else fill_potential)
            for item in c["items"]:
                ws_cls.cell(row=r, column=1, value=c["id"]).fill = fill_bg
                ws_cls.cell(row=r, column=2, value=c["match_type"]).fill = fill_bg
                ws_cls.cell(row=r, column=3, value=f"{c['confidence']}%").fill = fill_bg
                ws_cls.cell(row=r, column=4, value=item.get("source_file", "")).fill = fill_bg
                ws_cls.cell(row=r, column=5, value=item.get("date_str", "")).fill = fill_bg
                ws_cls.cell(row=r, column=6, value=item.get("raw_narration", "")).fill = fill_bg
                ws_cls.cell(row=r, column=7, value=item.get("debit", 0.0)).fill = fill_bg
                ws_cls.cell(row=r, column=8, value=item.get("credit", 0.0)).fill = fill_bg
                ws_cls.cell(row=r, column=9, value=item.get("balance", "")).fill = fill_bg
                ws_cls.cell(row=r, column=10, value=item.get("ref_no", "")).fill = fill_bg
                ws_cls.cell(row=r, column=11, value=c.get("reason", "")).fill = fill_bg

                for c_i in range(1, 12):
                    ws_cls.cell(row=r, column=c_i).border = thin_border
                r += 1

        # Auto-adjust column widths
        for sheet in [ws_sum, ws_cls]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        return output_path
